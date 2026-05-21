"""Smoke tests upload xlsx de vendas manuais."""
import io
from datetime import date, timedelta


def _gerar_xlsx_em_memoria(linhas):
    """Helper: cria xlsx in-memory com [linhas] = [(data, produto, qtd)]."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(['Data', 'Produto', 'Quantidade'])
    for data_, prod, qtd in linhas:
        ws.append([data_, prod, qtd])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_template_xlsx_inclui_receitas(app, loja, catalogo):
    """Template gera xlsx valido com header + linhas de exemplo."""
    from openpyxl import load_workbook

    from app.services.vendas_manuais import gerar_template_xlsx
    blob = gerar_template_xlsx(loja)
    wb = load_workbook(io.BytesIO(blob))
    ws = wb['Vendas']
    headers = [c.value for c in ws[1]]
    assert headers == ['Data', 'Produto', 'Quantidade']
    # Linha 2 tem dado (uma receita do catalogo)
    assert ws.cell(row=2, column=2).value == 'Croissant Tradicional'
    # Aba de instrucoes existe
    assert 'Como usar' in wb.sheetnames


def test_upload_xlsx_grava_multiplas_datas(app, admin_user, loja, catalogo):
    """Upload com 3 datas distintas cria 3 VendaManualLoja."""
    from app.models import VendaManualLoja
    from app.services.vendas_manuais import aplicar_vendas_xlsx, parsear_xlsx

    hoje_ = date.today()
    buf = _gerar_xlsx_em_memoria([
        (hoje_, 'Croissant Tradicional', 5),
        (hoje_ - timedelta(days=1), 'Croissant Tradicional', 7),
        (hoje_ - timedelta(days=2), 'Croissant Tradicional', 10),
    ])
    parseados = parsear_xlsx(buf, loja.id)
    assert len(parseados) == 3
    resultado = aplicar_vendas_xlsx(parseados, loja.id, admin_user)
    assert len(resultado['aplicados']) == 3
    assert len(resultado['datas_unicas']) == 3

    vendas = VendaManualLoja.query.filter_by(loja_id=loja.id).order_by(
        VendaManualLoja.data_venda).all()
    assert len(vendas) == 3
    assert vendas[0].quantidade == 10
    assert vendas[-1].quantidade == 5  # mais recente


def test_upload_xlsx_aceita_data_dd_mm_yyyy(app, admin_user, loja, catalogo):
    """Parser aceita data como string DD/MM/YYYY."""
    from app.services.vendas_manuais import parsear_xlsx

    buf = _gerar_xlsx_em_memoria([
        ('15/04/2026', 'Croissant Tradicional', 5),
        ('2026-04-16', 'Croissant Tradicional', 8),  # YYYY-MM-DD tb funciona
    ])
    parseados = parsear_xlsx(buf, loja.id)
    erros = [p for p in parseados if p.get('erro')]
    assert not erros, f'erros inesperados: {erros}'
    assert parseados[0]['data_venda'] == date(2026, 4, 15)
    assert parseados[1]['data_venda'] == date(2026, 4, 16)


def test_upload_xlsx_ignora_linhas_invalidas(app, admin_user, loja, catalogo):
    """Linhas com data ruim, qtd zero, nome vazio sao ignoradas
    sem bloquear linhas validas."""
    from app.services.vendas_manuais import aplicar_vendas_xlsx, parsear_xlsx

    buf = _gerar_xlsx_em_memoria([
        (date.today(), 'Croissant Tradicional', 5),  # OK
        ('texto-invalido', 'Croissant', 3),  # data ruim
        (date.today(), '', 7),  # nome vazio
        (date.today(), 'Croissant', 0),  # qtd zero — silenciosa
        (date.today(), 'Produto inexistente xyz', 4),  # nao resolve
    ])
    parseados = parsear_xlsx(buf, loja.id)
    resultado = aplicar_vendas_xlsx(parseados, loja.id, admin_user)
    assert len(resultado['aplicados']) == 1
    assert len(resultado['ignorados']) >= 3  # 3 explicitos + qtd=0 sumiu antes

"""Export XLSX da aba Produtos do /entregas (08/08/2026, pedido do dono).

Rota `/entregas/produtos.xlsx` re-agrega no SERVIDOR pelo MESMO motor da aba
(`_produtos_do_dia`) — nunca o estado do navegador. Contrato ajustado pelo
dono no 1º uso real (08/08): UMA aba só, "A produzir" ABAIXO de "Vendidos"
(a 2ª aba passava despercebida no celular) e SEM valores em R$ (a planilha
circula com a equipe; "ninguém precisa saber dos valores, só eu").
"""
import io

from openpyxl import load_workbook

from app.extensions import db
from app.models import (
    PedidoOnline,
    PedidoOnlineItem,
    Produto,
    ProdutoItem,
    Receita,
    Usuario,
)
from app.utils import hoje


def _admin_client(app):
    u = Usuario(nome='Adm', login='adm_xlsx', papel='admin')
    u.set_senha('x' * 8)
    db.session.add(u)
    db.session.commit()
    c = app.test_client()
    with c.session_transaction() as s:
        s['_user_id'] = str(u.id)
        s['_fresh'] = True
    return c


def _cenario(janela='06:00–10:00'):
    """Cesta (2 croissants por unidade) vendida 3x."""
    r = Receita(nome='Croissant Tradicional', categoria='Croissants',
                rendimento_qtd=1, rendimento_unidade='un', peso_base=100)
    db.session.add(r)
    db.session.flush()
    cesta = Produto(nome='Cesta Dia dos Pais', ativo=True)
    db.session.add(cesta)
    db.session.flush()
    db.session.add(ProdutoItem(produto_id=cesta.id, tipo='receita',
                               receita_id=r.id, item_nome=r.nome,
                               quantidade=2))
    p = PedidoOnline(codigo='XLSX1', nome_cliente='C', email_cliente='x@x.com',
                     status='pago', modo_entrega='agendada',
                     data_entrega=hoje(), janela_entrega=janela,
                     valor_total=450)
    db.session.add(p)
    db.session.flush()
    db.session.add(PedidoOnlineItem(pedido_id=p.id, kind='produto',
                                    produto_id=cesta.id, nome=cesta.nome,
                                    preco_unitario=150, quantidade=3,
                                    subtotal=450))
    db.session.commit()
    return r, cesta


def _linhas(resp):
    wb = load_workbook(io.BytesIO(resp.data), read_only=True)
    ws = wb[wb.sheetnames[0]]
    return wb, [[c_.value for c_ in row] for row in ws.rows]


def test_uma_aba_so_com_a_produzir_abaixo_dos_vendidos(app):
    with app.app_context():
        _cenario()
        c = _admin_client(app)
        resp = c.get(f'/entregas/produtos.xlsx?data={hoje().isoformat()}')
        assert resp.status_code == 200
        assert 'spreadsheetml' in resp.mimetype
        wb, linhas = _linhas(resp)
        # UMA aba só (a 2ª passava despercebida no celular — dono 08/08).
        assert wb.sheetnames == ['Produtos do dia']
        col_a = [li[0] for li in linhas if li and li[0]]
        # Vendidos primeiro, A produzir ABAIXO na mesma aba:
        assert col_a.index('Vendidos no dia') < col_a.index(
            'A produzir (cestas explodidas)')

        vendida = next(li for li in linhas if li[0] == 'Cesta Dia dos Pais')
        assert vendida[1] == 3                        # como vendida
        explodida = next(li for li in linhas
                         if li[0] == 'Croissant Tradicional')
        assert explodida[1] == 6 and explodida[2] == 'un'   # 3 x 2
        assert 'Cesta Dia dos Pais' in (explodida[3] or '')
        totais = [li for li in linhas if li[0] == 'TOTAL']
        assert totais[0][1] == 3                      # vendidos
        assert '6 un' in str(totais[1][1])            # produção por unidade


def test_sem_nenhum_valor_em_reais(app):
    """"Ninguém precisa saber dos valores, só eu" — a planilha circula com
    a equipe de montagem: nenhum preço/total em R$ no arquivo."""
    with app.app_context():
        _cenario()
        c = _admin_client(app)
        resp = c.get(f'/entregas/produtos.xlsx?data={hoje().isoformat()}')
        _wb, linhas = _linhas(resp)
        achatado = [str(v) for li in linhas for v in li if v is not None]
        assert not any('R$' in v or 'Preço' in v or 'Total (R$)' in v
                       for v in achatado)
        # O valor da venda (450) não aparece em lugar nenhum:
        assert not any(v in ('450', '450.0', '150', '150.0')
                       for v in achatado)


def test_filtro_de_janela_vale_no_xlsx(app):
    with app.app_context():
        _cenario(janela='06:00–10:00')
        c = _admin_client(app)
        resp = c.get(f'/entregas/produtos.xlsx?data={hoje().isoformat()}'
                     '&janela=09:00–10:00')
        _wb, linhas = _linhas(resp)
        assert not any(li[0] == 'Cesta Dia dos Pais' for li in linhas)
        assert any('09:00–10:00' in str(li[0]) for li in linhas
                   if li and li[0])


def test_xlsx_exige_login(app):
    r = app.test_client().get('/entregas/produtos.xlsx')
    assert r.status_code in (302, 401, 403)

"""Plano de carreira: importa proposta e vínculos sem alterar a folha."""
import base64
import hashlib
import io

from openpyxl import Workbook

from app.extensions import db


def _xlsx(decisao=''):
    wb = Workbook()
    wb.remove(wb.active)
    guia = wb.create_sheet('Guia e Controle')
    guia.append(['PLANO']); guia.append(['revisão técnica: teste']); guia.append([])
    guia.append(['#', 'Validação', 'Motivo', 'Responsável', 'Estado', 'Evidência esperada', 'Bloqueia implantação?'])
    guia.append([1, 'Validar CCT', 'Evitar erro', 'DP', 'Pendente', 'Parecer', 'Sim'])
    plano = wb.create_sheet('Plano N1-N5')
    plano.append(['Chave', 'Família', 'Nível nº', 'Cargo proposto', 'Papel no nível', 'Unidade', 'Base de referência', 'Multiplicador', 'Salário do nível', 'Horas/mês', 'Equiv. mensal', 'Complemento de função (R$)', 'Total alvo mensal', 'Vídeos mínimos acumulados', 'Tempo mínimo anterior (meses)', 'Checklist mínimo', 'Certificação prática mínima', 'Observação'])
    plano.append(['Atendimento|1', 'Atendimento', 1, 'Atendente 1', 'Entrada', 'R$/mês', 2130.4, 1, 2130.4, 0, 2130.4, 0, 2130.4, 1, 0, .8, 'Atender sem erro', 'Teste'])
    matriz = wb.create_sheet('Matriz Prática')
    matriz.append(['Chave', 'Família', 'Nível nº', 'Nível', 'Certificação prática mínima', 'Forma de avaliação', 'Regra'])
    matriz.append(['Atendimento|1', 'Atendimento', 1, 'N1', 'Recepção sem erro', 'Líder', 'Obrigatória'])
    regras = wb.create_sheet('Regras Promoção')
    regras.append(['Transição', 'Tempo mínimo no nível atual', 'Conteúdo mínimo', 'Checklist prático', 'Certificação prática', 'Evidência adicional', 'Aprovação', 'Se não atingir'])
    regras.append(['N1 → N2', '6 meses', '100%', '≥ 85%', 'N2 concluída', '2 ciclos', 'Líder + gerente', 'Reavaliar'])
    conteudos = wb.create_sheet('Conteúdos Mínimos')
    conteudos.append(['Família', 'Código', 'Módulo', 'Título do vídeo', 'Categoria', 'Obrigatório a partir do nível', 'Nível', 'Objetivo do conteúdo'])
    conteudos.append(['Atendimento', 'M01.01', 'Módulo 1 — Cultura', 'Nossa história', 'Cultura', 1, 'N1', 'Conhecer'])
    enquad = wb.create_sheet('Enquadramento Atual')
    enquad.append(['Nome', 'Cargo atual', 'Salário base atual', 'Complementos/HE/DSR atuais', 'Total atual', 'Família sugerida', 'Nível inicial', 'Chave', 'Cargo proposto', 'Base alvo mensal', 'Complemento função alvo', 'Total alvo', 'Dif. base', 'Dif. total', 'Status do cenário total', 'Nota de transição', 'Fonte / competência do cenário atual', 'Decisão registrada'])
    enquad.append(['Amânda de Souza', 'ATENDENTE', 2000, 0, 2000, 'Atendimento', 1, 'Atendimento|1', 'Atendente 1', 2130.4, 0, 2130.4, 130.4, 130.4, 'Abaixo da referência sugerida', 'Avaliar', 'Holerite teste', decisao])
    out = io.BytesIO(); wb.save(out); return out.getvalue()


def _funcionario():
    from app.models import Funcionario
    f = Funcionario(nome='AMANDA DE SOUZA', cpf='42326669827', funcao='ATENDENTE', salario_base=2000, ativo=True)
    db.session.add(f); db.session.commit(); return f


def _video():
    from app.models import TreinoTrilha, TreinoVideo
    trilha = TreinoTrilha(nome='Módulo 1 — Cultura', ordem=1)
    db.session.add(trilha); db.session.flush()
    video = TreinoVideo(trilha_id=trilha.id, titulo='Nossa história', ordem=1)
    db.session.add(video); db.session.commit(); return video


def test_previa_vincula_nome_e_video_sem_gravar(app):
    from app.models import Cargo, PlanoCarreiraFaixa
    from app.services import plano_carreira_import as svc
    _funcionario(); _video(); dados = svc.prever(_xlsx())
    assert dados['resumo']['pessoas_vinculadas'] == 1
    assert dados['resumo']['videos_vinculados'] == 1
    assert dados['resumo']['cargos_a_criar'] == 1
    assert dados['resumo']['aprovacoes_a_aplicar'] == 0
    assert dados['faixas'][0]['certificacao_pratica'] == 'Recepção sem erro'
    assert PlanoCarreiraFaixa.query.count() == 0
    assert Cargo.query.count() == 0


def test_aplicar_vincula_faixa_e_aplica_cargo_aprovado(app):
    from app.models import (
        Cargo,
        PlanoCarreiraCargoVinculo,
        PlanoCarreiraConteudo,
        PlanoCarreiraEnquadramento,
    )
    from app.services import plano_carreira_import as svc
    f = _funcionario(); video = _video(); svc.aplicar(_xlsx('Aprovado'), 'plano.xlsx')
    db.session.expire_all()
    f = db.session.get(type(f), f.id)
    e = PlanoCarreiraEnquadramento.query.one()
    assert e.funcionario_id == f.id and e.decisao == 'Aprovado'
    assert PlanoCarreiraConteudo.query.one().treino_video_id == video.id
    assert PlanoCarreiraCargoVinculo.query.count() == 1
    assert Cargo.query.count() == 1
    assert f.cargo.nome == 'Atendente 1'
    assert f.salario_base == 2130.4 and f.funcao == 'Atendente 1'


def test_proposta_nao_altera_cargo_real(app):
    from app.services import plano_carreira_import as svc
    f = _funcionario()

    svc.aplicar(_xlsx('Proposta final'), 'plano.xlsx')

    f = db.session.get(type(f), f.id)
    assert f.cargo_id is None
    assert f.salario_base == 2000 and f.funcao == 'ATENDENTE'


def test_trocar_para_cargo_do_plano_sincroniza_enquadramento(app):
    from app.models import Cargo, PlanoCarreiraEnquadramento
    from app.services import plano_carreira_import as svc
    f = _funcionario()
    svc.aplicar(_xlsx(), 'plano.xlsx')
    f = db.session.get(type(f), f.id)
    f.cargo = Cargo.query.filter_by(nome='Atendente 1').one()

    faixa = svc.sincronizar_enquadramento_com_cargo(f)
    db.session.commit()

    e = PlanoCarreiraEnquadramento.query.one()
    assert faixa.nivel == 1
    assert e.familia == 'Atendimento' and e.nivel == 1
    assert e.cargo_proposto == 'Atendente 1'
    assert e.decisao == 'Aprovado'


def test_reimportacao_preserva_decisao_manual_quando_planilha_vazia(app):
    from app.models import PlanoCarreiraEnquadramento
    from app.services import plano_carreira_import as svc
    _funcionario(); svc.aplicar(_xlsx(), 'primeiro.xlsx')
    e = PlanoCarreiraEnquadramento.query.one(); e.decisao = 'Proposta final'; db.session.commit()
    svc.aplicar(_xlsx(), 'segundo.xlsx')
    assert PlanoCarreiraEnquadramento.query.one().decisao == 'Proposta final'


def _login_owner(app):
    from app.models import Usuario
    u = Usuario(nome='Dono', login='dono_carreira', papel='admin', is_owner=True); u.set_senha('senha123')
    db.session.add(u); db.session.commit(); c = app.test_client()
    with c.session_transaction() as sess: sess['_user_id'] = str(u.id); sess['_fresh'] = True
    return c


def test_fluxo_da_tela_importa_e_exibe_plano(app):
    from app.models import PlanoCarreiraEnquadramento
    _funcionario(); c = _login_owner(app); raw = _xlsx()
    resposta = c.post('/rh/plano-carreira/importar', data={'arquivo': (io.BytesIO(raw), 'plano.xlsx')}, content_type='multipart/form-data')
    assert resposta.status_code == 200 and 'Prévia concluída' in resposta.data.decode()
    assert 'cargo(s) serão criados' in resposta.data.decode()
    resposta = c.post('/rh/plano-carreira/importar/aplicar', data={'arquivo_b64': base64.b64encode(raw).decode('ascii'), 'arquivo_sha': hashlib.sha256(raw).hexdigest(), 'arquivo_nome': 'plano.xlsx'}, follow_redirects=True)
    assert resposta.status_code == 200 and 'Plano vinculado' in resposta.data.decode()
    assert PlanoCarreiraEnquadramento.query.count() == 1


def test_salvar_decisao_aprovada_aplica_cargo_na_ficha(app):
    from app.models import PlanoCarreiraEnquadramento
    from app.services import plano_carreira_import as svc
    f = _funcionario()
    svc.aplicar(_xlsx(), 'plano.xlsx')
    e = PlanoCarreiraEnquadramento.query.one()
    c = _login_owner(app)

    resposta = c.post(
        f'/rh/plano-carreira/enquadramentos/{e.id}/decisao',
        data={'decisao': 'Aprovado', 'voltar': '/rh/plano-carreira'},
        follow_redirects=True,
    )

    assert resposta.status_code == 200
    assert 'cargo vinculado' in resposta.data.decode()
    f = db.session.get(type(f), f.id)
    assert f.cargo.nome == 'Atendente 1'


def test_admin_comum_nao_acessa_plano(app, admin_user):
    c = app.test_client()
    with c.session_transaction() as sess: sess['_user_id'] = str(admin_user.id); sess['_fresh'] = True
    assert c.get('/rh/plano-carreira').status_code == 403

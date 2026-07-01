"""Planilha de contagem em branco (balanço de loja, 'caminho ao contrário'):
lista todos os itens que a loja pede + produtos ativos, coluna Quantidade vazia.
"""
import io


def _receita(nome, categoria='Paes', sugerir=True, arquivada=False):
    from datetime import datetime

    from app.extensions import db
    from app.models import Receita
    r = Receita(nome=nome, categoria=categoria, rendimento_qtd=1,
                rendimento_unidade='un', peso_base=100.0,
                sugerir_pedido_loja=sugerir,
                arquivada_em=datetime(2026, 1, 1) if arquivada else None)
    db.session.add(r)
    db.session.commit()
    return r


def _produto(nome, ativo=True):
    from app.extensions import db
    from app.models import Produto
    p = Produto(nome=nome, ativo=ativo)
    db.session.add(p)
    db.session.commit()
    return p


def _linhas(blob):
    """Devolve {nome: (categoria, quantidade)} das linhas de dados do xlsx."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob))
    ws = wb['Contagem']
    out = {}
    for row in ws.iter_rows(min_row=6, values_only=True):
        cat, nome, qtd = row[0], row[1], row[2]
        if nome:
            out[nome] = (cat, qtd)
    return out


def test_template_lista_so_o_que_a_loja_pede(app):
    from app.services.estoque_loja_lote import gerar_xlsx_template_balanco
    _receita('Croissant Tradicional', categoria='Croissants')
    _receita('Massa para folhar', categoria='Insumos', sugerir=False)  # não pede
    _receita('Pão Arquivado', arquivada=True)                          # fora
    _produto('Granola 500g', ativo=True)
    _produto('Produto Morto', ativo=False)                             # fora

    linhas = _linhas(gerar_xlsx_template_balanco())
    assert 'Croissant Tradicional' in linhas
    assert 'Granola 500g' in linhas                     # produto ativo entra
    assert linhas['Granola 500g'][0] == 'Produtos'      # seção Produtos
    assert 'Massa para folhar' not in linhas            # sugerir_pedido_loja=False
    assert 'Pão Arquivado' not in linhas                # arquivada
    assert 'Produto Morto' not in linhas                # produto inativo


def test_template_quantidade_vem_em_branco(app):
    from app.services.estoque_loja_lote import gerar_xlsx_template_balanco
    _receita('Croissant Tradicional')
    linhas = _linhas(gerar_xlsx_template_balanco())
    assert linhas['Croissant Tradicional'][1] is None   # coluna Quantidade vazia


def test_template_ordena_por_categoria_e_nome(app):
    from app.services.estoque_loja_lote import gerar_xlsx_template_balanco
    _receita('Baguete', categoria='Paes')
    _receita('Croissant', categoria='Croissants')
    _receita('Amendoa', categoria='Croissants')
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(gerar_xlsx_template_balanco()))
    ws = wb['Contagem']
    ordem = [(r[0], r[1]) for r in ws.iter_rows(min_row=6, values_only=True) if r[1]]
    # Croissants (Amendoa, Croissant) antes de Paes (Baguete)
    assert ordem == [('Croissants', 'Amendoa'), ('Croissants', 'Croissant'),
                     ('Paes', 'Baguete')]


def test_rota_baixa_xlsx(app, admin_user):
    _receita('Croissant Tradicional')
    c = app.test_client()
    c.post('/auth/login', data={'login': admin_user.login, 'senha': '123'},
           follow_redirects=True)
    resp = c.get('/pedidos/estoque-loja/balanco-template.xlsx')
    assert resp.status_code == 200
    assert 'spreadsheetml' in resp.headers['Content-Type']
    assert b'PK' == resp.get_data()[:2]                 # zip/xlsx magic


def test_rota_exige_login(app):
    c = app.test_client()
    resp = c.get('/pedidos/estoque-loja/balanco-template.xlsx')
    assert resp.status_code in (301, 302, 403)

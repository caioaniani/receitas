"""Exportação XLSX da tabela de preços + preço interno nas fichas (02/07/2026).

- GET /receitas/precos.xlsx (owner): PRODUTO | CUSTO | PRECO LOJA | PRECO SITE
  | PRECO INTERNO | ATACADO (+ TIPO/CATEGORIA), receitas + produtos + cestas.
- A ficha da receita e o form do produto agora salvam preco_interno (antes o
  campo só existia na tela /receitas/precos — reportado pelo dono).
"""
import io

from openpyxl import load_workbook

from app.extensions import db
from app.models import Produto, Receita


def _receita(nome='Croissant', **kw):
    r = Receita(nome=nome, categoria='Paes', rendimento_qtd=1,
                rendimento_unidade='un', peso_base=100.0, **kw)
    db.session.add(r)
    db.session.commit()
    return r


def _login(app, user):
    client = app.test_client()
    client.post('/auth/login', data={'login': user.login, 'senha': '123'})
    return client


def test_export_xlsx_layout_e_valores(app, owner_user):
    r = _receita('Sourdough', preco_loja=25.0, preco_site=32.0,
                 preco_interno=15.0, preco_venda=20.0)
    p = Produto(nome='Cesta Café', categoria='Cestas', ativo=True,
                preco_loja=100.0, preco_site=120.0, preco_interno=80.0,
                preco_atacado=90.0)
    db.session.add(p)
    db.session.commit()

    client = _login(app, owner_user)
    resp = client.get('/receitas/precos.xlsx')
    assert resp.status_code == 200
    assert 'spreadsheetml' in resp.content_type

    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb['Precos']
    cabecalho = [c.value for c in ws[4]]
    assert cabecalho == ['PRODUTO', 'CUSTO', 'PRECO LOJA', 'PRECO SITE',
                         'PRECO INTERNO', 'ATACADO', 'TIPO', 'CATEGORIA']
    linhas = {row[0]: row for row in ws.iter_rows(min_row=5, values_only=True)}
    assert linhas['Sourdough'][2:6] == (25.0, 32.0, 15.0, 20.0)
    assert linhas['Sourdough'][6] == 'Receita'
    assert linhas['Cesta Café'][2:6] == (100.0, 120.0, 80.0, 90.0)
    _ = r  # receita usada via planilha


def test_export_exige_owner(app, admin_user):
    client = _login(app, admin_user)
    resp = client.get('/receitas/precos.xlsx')
    assert resp.status_code in (302, 403)   # admin comum não baixa


def test_ficha_salva_preco_interno(app, admin_user):
    r = _receita('Baguete')
    client = _login(app, admin_user)
    resp = client.post(f'/receitas/{r.id}/salvar', data={
        'nome': 'Baguete', 'categoria': 'Paes',
        'rendimento_qtd': '1', 'rendimento_unidade': 'un',
        'peso_base': '100', 'preco_venda': '10,00', 'preco_loja': '12,00',
        'preco_site': '14,00', 'preco_interno': '8,50',
    })
    assert resp.status_code in (200, 302)
    db.session.refresh(r)
    assert r.preco_interno == 8.5


def test_produto_salva_preco_interno(app, admin_user):
    p = Produto(nome='Kit Lanche', categoria='Kits', ativo=True)
    db.session.add(p)
    db.session.commit()
    client = _login(app, admin_user)
    resp = client.post(f'/produtos/{p.id}/salvar', data={
        'nome': 'Kit Lanche', 'categoria': 'Kits',
        'preco_atacado': '30,00', 'preco_loja': '35,00',
        'preco_site': '40,00', 'preco_interno': '25,00',
    })
    assert resp.status_code in (200, 302)
    db.session.refresh(p)
    assert p.preco_interno == 25.0

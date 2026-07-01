"""Itens vendidos SEPARADOS POR LOJA + export XLSX (uma aba por loja +
Consolidado). A API do Seru e mockada — sem rede."""
import io
from unittest.mock import patch

from app.services import vendas_itens

DIA = '2026-06-15'


def _pedido(pid, loja, itens, created=f'{DIA}T13:00:00Z'):
    return {'id': pid, 'createdAt': created, 'canceledAt': None,
            'company': {'name': loja},
            'items': [{'name': n, 'quantity': q, 'total': t} for n, q, t in itens]}


PEDIDOS = [
    _pedido(1, 'Ribeiro do Vale', [('Cookie', 5, 50.0), ('Brioche', 2, 20.0)]),
    _pedido(2, 'Ribeiro do Vale', [('Cookie', 3, 30.0)]),
    _pedido(3, 'Nebraska', [('Cookie', 1, 10.0)]),
]


def _agg(app):
    from datetime import date
    with patch('app.services.seru.listar_pedidos_completo', return_value=PEDIDOS):
        return vendas_itens.agregar_itens_por_loja(date(2026, 6, 15), date(2026, 6, 15))


def test_agrega_separado_por_loja(app):
    d = _agg(app)
    lojas = {lo['loja']: lo for lo in d['lojas']}
    assert set(lojas) == {'Ribeiro do Vale', 'Nebraska'}
    # Ribeiro: Cookie 5+3=8 / R$80, Brioche 2 / R$20; 2 pedidos
    rib = lojas['Ribeiro do Vale']
    assert rib['total_pedidos'] == 2
    assert rib['faturamento'] == 100.0
    cookie = next(p for p in rib['produtos'] if p['nome'] == 'Cookie')
    assert cookie['qtd'] == 8 and cookie['faturamento'] == 80.0
    assert cookie['n_pedidos'] == 2
    # Nebraska: Cookie 1 / R$10
    assert lojas['Nebraska']['faturamento'] == 10.0
    # Consolidado soma as duas: Cookie 9 / R$90, Brioche 2 / R$20
    cons = {p['nome']: p for p in d['consolidado']}
    assert cons['Cookie']['qtd'] == 9 and cons['Cookie']['faturamento'] == 90.0
    assert d['faturamento_total'] == 110.0
    assert d['total_pedidos'] == 3


def test_xlsx_uma_aba_por_loja_mais_consolidado(app):
    from openpyxl import load_workbook
    d = _agg(app)
    blob = vendas_itens.gerar_xlsx_itens_por_loja(d)
    wb = load_workbook(io.BytesIO(blob))
    # Consolidado primeiro, depois uma aba por loja (ordem alfabetica)
    assert wb.sheetnames[0] == 'Consolidado'
    assert 'Nebraska' in wb.sheetnames
    assert 'Ribeiro do Vale' in wb.sheetnames
    # A aba da loja tem o cabecalho e ao menos 1 produto
    ws = wb['Ribeiro do Vale']
    assert ws['A4'].value == 'Produto'
    nomes = {ws.cell(row=r, column=1).value for r in range(5, 8)}
    assert 'Cookie' in nomes


def test_rota_xlsx_baixa_arquivo(app, admin_user):
    c = app.test_client()
    c.post('/auth/login', data={'login': admin_user.login, 'senha': '123'},
           follow_redirects=True)
    with patch('app.services.seru.listar_pedidos_completo', return_value=PEDIDOS):
        r = c.get(f'/pdv/itens-vendidos.xlsx?inicio={DIA}&fim={DIA}')
    assert r.status_code == 200
    assert 'spreadsheetml.sheet' in r.headers['Content-Type']
    assert '.xlsx' in r.headers.get('Content-Disposition', '')
    # e um xlsx de verdade (magic zip PK)
    assert r.data[:2] == b'PK'


def test_rota_por_loja_json(app, admin_user):
    c = app.test_client()
    c.post('/auth/login', data={'login': admin_user.login, 'senha': '123'},
           follow_redirects=True)
    with patch('app.services.seru.listar_pedidos_completo', return_value=PEDIDOS):
        r = c.get(f'/pdv/api/itens-vendidos-por-loja?inicio={DIA}&fim={DIA}')
    assert r.status_code == 200
    j = r.get_json()
    assert j['ok'] is True
    assert len(j['lojas']) == 2
    assert j['faturamento_total'] == 110.0

"""Importa o PLANO DE CONTEÚDO do treinamento por planilha (13/08/2026).

O dono montou a "Universidade" numa planilha: 9 módulos, 140 aulas, cada uma
com roteiro de gravação completo (objetivo, gancho, demonstração,
comportamento esperado, desafio da semana). Os vídeos ainda não existem —
o que entra aqui é a ESTRUTURA: módulo vira `TreinoTrilha`, aula vira
`TreinoVideo` RASCUNHO (`ativo=False`) com o roteiro anexado. Quem grava
abre a aula no admin, lê o roteiro ali e sobe o arquivo no fluxo que já
existe (upload direto pro Cloudflare na tela da aula).

Regras de peso:
- TUDO nasce desativado (trilha e aula): o funcionário não vê 9 trilhas
  vazias — cada módulo é ativado quando os vídeos dele estiverem no ar;
- idempotente: re-importar uma planilha revisada ATUALIZA os roteiros sem
  duplicar nada (match: trilha pelo nome, aula pelo Nº dentro da trilha);
- aula que JÁ TEM vídeo gravado nunca tem título/duração sobrescritos
  (produção no ar não muda por planilha) — só o roteiro acompanha revisão;
- o import nunca DESATIVA nem apaga nada.
"""
import io
import logging
import re
import unicodedata

from app.extensions import db
from app.models import TreinoTrilha, TreinoVideo

logger = logging.getLogger(__name__)


def _norm(s):
    s = unicodedata.normalize('NFKD', str(s or ''))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return ' '.join(s.upper().split())


def _montar_roteiro(d):
    """Texto único da aula, com as seções nomeadas — é o que aparece no card
    "Roteiro de gravação" da tela da aula."""
    partes = []
    if d.get('Código'):
        partes.append(f"[{d['Código']}] Público: {d.get('Público') or 'Todos'}")
    for rotulo, col in (('OBJETIVO', 'Objetivo'),
                        ('ROTEIRO', 'Roteiro do vídeo'),
                        ('DEMONSTRAÇÃO / EXEMPLO', 'Demonstração / exemplo'),
                        ('COMPORTAMENTO ESPERADO', 'Comportamento esperado'),
                        ('DESAFIO DA SEMANA', 'Desafio da semana'),
                        ('OBSERVAÇÕES', 'Observações')):
        txt = (d.get(col) or '').strip()
        if txt:
            partes.append(f'{rotulo}\n{txt}')
    return '\n\n'.join(partes)


def _num_modulo(nome_modulo):
    m = re.search(r'(\d+)', nome_modulo or '')
    return int(m.group(1)) if m else 0


def _minutos(bruto):
    try:
        return max(0.0, float(str(bruto).replace(',', '.')))
    except (TypeError, ValueError):
        return 0.0


def ler_planilha(conteudo_bytes):
    """Lê a aba de roteiros. Devolve (linhas, avisos). Linha sem código ou
    sem título vira AVISO — nunca some em silêncio."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(conteudo_bytes), data_only=True)
    ws = None
    for cand in wb.worksheets:
        cab = [str(c.value or '').strip() for c in next(cand.iter_rows(max_row=1))]
        if any('Módulo' in c for c in cab) and \
                any('Título' in c for c in cab):
            ws, cabecalho = cand, cab
            break
    if ws is None:
        raise ValueError('Não achei a aba de roteiros (procuro colunas '
                         '"Módulo" e "Título do vídeo").')
    linhas, avisos = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(cabecalho,
                     [str(v).strip() if v is not None else '' for v in row]))
        if not any(d.values()):
            continue
        titulo = d.get('Título do vídeo') or ''
        modulo = d.get('Módulo') or ''
        try:
            aula_n = int(float(d.get('Nº da aula') or 0))
        except (TypeError, ValueError):
            aula_n = 0
        if not titulo or not modulo or aula_n <= 0:
            avisos.append(f"linha \"{d.get('Código') or titulo or '?'}\": sem "
                          'módulo, título ou nº da aula — IGNORADA.')
            continue
        linhas.append({'modulo': modulo, 'ordem': aula_n,
                       'titulo': titulo[:200],
                       'minutos': _minutos(d.get('Duração sugerida (min)')),
                       'publico': d.get('Público') or 'Todos',
                       'roteiro': _montar_roteiro(d)})
    if not linhas:
        raise ValueError('Nenhuma aula legível na planilha.')
    return linhas, avisos


def importar(conteudo_bytes):
    """Cria/atualiza trilhas e aulas a partir da planilha. Devolve stats."""
    linhas, avisos = ler_planilha(conteudo_bytes)
    stats = {'trilhas_criadas': 0, 'aulas_criadas': 0,
             'roteiros_atualizados': 0, 'aulas_com_video_preservadas': 0,
             'avisos': avisos}

    por_nome = {_norm(t.nome): t for t in TreinoTrilha.query.all()}
    modulos = {}
    for ln in linhas:
        modulos.setdefault(ln['modulo'], []).append(ln)

    maior_ordem = (db.session.query(db.func.max(TreinoTrilha.ordem))
                   .scalar() or 0)
    for nome_mod, aulas in modulos.items():
        t = por_nome.get(_norm(nome_mod))
        if t is None:
            maior_ordem += 1
            t = TreinoTrilha(
                nome=nome_mod[:150],
                descricao=f"Público: {aulas[0]['publico']} · plano de "
                          f'conteúdo importado por planilha',
                ordem=_num_modulo(nome_mod) or maior_ordem,
                ativa=False)
            db.session.add(t)
            db.session.flush()
            por_nome[_norm(nome_mod)] = t
            stats['trilhas_criadas'] += 1
        # Carga horária acompanha a planilha (minutos sugeridos somados).
        t.carga_horaria_minutos = int(round(
            sum(a['minutos'] for a in aulas)))

        existentes = {v.ordem: v for v in t.videos}
        for a in aulas:
            v = existentes.get(a['ordem'])
            if v is None:
                db.session.add(TreinoVideo(
                    trilha_id=t.id, titulo=a['titulo'], ordem=a['ordem'],
                    duracao_segundos=int(a['minutos'] * 60),
                    ativo=False, roteiro=a['roteiro']))
                stats['aulas_criadas'] += 1
                continue
            if (v.roteiro or '') != a['roteiro']:
                v.roteiro = a['roteiro']
                stats['roteiros_atualizados'] += 1
            if v.video_externo_id:
                # Vídeo gravado: título/duração são produção no ar — a
                # planilha não os sobrescreve (a duração real veio do
                # Cloudflare).
                stats['aulas_com_video_preservadas'] += 1
            else:
                v.titulo = a['titulo']
                v.duracao_segundos = int(a['minutos'] * 60)
    db.session.commit()
    logger.info('treino_roteiros: import %s', stats)
    return stats

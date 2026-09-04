"""Importa o plano de cargos/carreira sem mexer na folha contratual."""
from __future__ import annotations

import hashlib
import io
import re
import unicodedata
from collections import defaultdict

from openpyxl import load_workbook

from app.extensions import db
from app.models import (
    Cargo,
    Funcionario,
    PlanoCarreiraCargoVinculo,
    PlanoCarreiraConteudo,
    PlanoCarreiraEnquadramento,
    PlanoCarreiraFaixa,
    PlanoCarreiraImportacao,
    PlanoCarreiraRegra,
    PlanoCarreiraValidacao,
    TreinoVideo,
)

MAX_BYTES = 4 * 1024 * 1024
DECISOES = ('Em avaliação', 'Aprovado', 'Proposta final')


class PlanoCarreiraErro(ValueError):
    pass


def _norm(valor):
    texto = unicodedata.normalize('NFKD', str(valor or ''))
    texto = ''.join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r'[^a-zA-Z0-9]+', ' ', texto).strip().casefold()
    return ' '.join(texto.split())


def _texto(valor, limite=None):
    texto = str(valor or '').strip()
    return texto[:limite] if limite else texto


def _numero(valor, default=0.0):
    if valor in (None, ''):
        return default
    try:
        return float(valor)
    except (TypeError, ValueError):
        return default


def _inteiro(valor, default=0):
    try:
        return int(float(valor))
    except (TypeError, ValueError):
        return default


def _aba(wb, nome):
    if nome not in wb.sheetnames:
        raise PlanoCarreiraErro(f'A planilha não contém a aba “{nome}”.')
    return wb[nome]


def _cabecalho(ws, primeira_coluna):
    alvo = _norm(primeira_coluna)
    for row in ws.iter_rows():
        valores = [cell.value for cell in row]
        if any(_norm(v) == alvo for v in valores):
            return row[0].row, {_norm(v): i for i, v in enumerate(valores)
                                if v not in (None, '')}
    raise PlanoCarreiraErro(
        f'Cabeçalho “{primeira_coluna}” não encontrado na aba “{ws.title}”.')


def _valor(row, colunas, *nomes):
    for nome in nomes:
        idx = colunas.get(_norm(nome))
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def _linhas(ws, primeira_coluna):
    cab, colunas = _cabecalho(ws, primeira_coluna)
    for row in ws.iter_rows(min_row=cab + 1, values_only=True):
        yield row, colunas


def ler(raw: bytes):
    if not raw:
        raise PlanoCarreiraErro('Selecione a planilha do plano de carreira.')
    if len(raw) > MAX_BYTES:
        raise PlanoCarreiraErro('A planilha ultrapassa o limite de 4 MB.')
    try:
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise PlanoCarreiraErro('O arquivo não é uma planilha xlsx válida.') from exc

    faixas = []
    for row, cols in _linhas(_aba(wb, 'Plano N1-N5'), 'Chave'):
        familia = _texto(_valor(row, cols, 'Família'), 100)
        nivel = _inteiro(_valor(row, cols, 'Nível nº'))
        cargo = _texto(_valor(row, cols, 'Cargo proposto'), 150)
        if not familia or not nivel or not cargo:
            continue
        faixas.append({
            'familia': familia, 'nivel': nivel, 'cargo_proposto': cargo,
            'papel': _texto(_valor(row, cols, 'Papel no nível')),
            'unidade': _texto(_valor(row, cols, 'Unidade'), 30),
            'base_referencia': _numero(_valor(row, cols, 'Base de referência')),
            'multiplicador': _numero(_valor(row, cols, 'Multiplicador'), 1),
            'salario_nivel': _numero(_valor(row, cols, 'Salário do nível')),
            'horas_mes': _numero(_valor(row, cols, 'Horas/mês')),
            'equivalente_mensal': _numero(_valor(row, cols, 'Equiv. mensal')),
            'complemento_funcao': _numero(_valor(row, cols, 'Complemento de função (R$)')),
            'total_alvo': _numero(_valor(row, cols, 'Total alvo mensal')),
            'videos_minimos': _inteiro(_valor(row, cols, 'Vídeos mínimos acumulados')),
            'tempo_minimo_meses': _inteiro(_valor(row, cols, 'Tempo mínimo anterior (meses)')),
            'checklist_minimo': _numero(_valor(row, cols, 'Checklist mínimo')),
            'certificacao_pratica': _texto(_valor(row, cols, 'Certificação prática mínima')),
            'observacao': _texto(_valor(row, cols, 'Observação')),
        })
    if not faixas:
        raise PlanoCarreiraErro('Nenhuma faixa válida foi encontrada em “Plano N1-N5”.')

    certificacoes = {}
    for row, cols in _linhas(_aba(wb, 'Matriz Prática'), 'Chave'):
        familia = _texto(_valor(row, cols, 'Família'), 100)
        nivel = _inteiro(_valor(row, cols, 'Nível nº'))
        if familia and nivel:
            certificacoes[(_norm(familia), nivel)] = _texto(
                _valor(row, cols, 'Certificação prática mínima'))
    for faixa in faixas:
        faixa['certificacao_pratica'] = certificacoes.get(
            (_norm(faixa['familia']), faixa['nivel']), faixa['certificacao_pratica'])

    regras = []
    for row, cols in _linhas(_aba(wb, 'Regras Promoção'), 'Transição'):
        transicao = _texto(_valor(row, cols, 'Transição'), 30)
        if not re.match(r'^N\d+\s*[→>-]+\s*N\d+$', transicao, re.I):
            continue
        regras.append({
            'transicao': transicao,
            'tempo_minimo': _texto(_valor(row, cols, 'Tempo mínimo no nível atual'), 150),
            'conteudo_minimo': _texto(_valor(row, cols, 'Conteúdo mínimo')),
            'checklist_minimo': _texto(_valor(row, cols, 'Checklist prático'), 50),
            'certificacao': _texto(_valor(row, cols, 'Certificação prática')),
            'evidencia': _texto(_valor(row, cols, 'Evidência adicional')),
            'aprovacao': _texto(_valor(row, cols, 'Aprovação')),
            'se_nao_atingir': _texto(_valor(row, cols, 'Se não atingir')),
        })

    conteudos = []
    for row, cols in _linhas(_aba(wb, 'Conteúdos Mínimos'), 'Família'):
        familia = _texto(_valor(row, cols, 'Família'), 100)
        codigo = _texto(_valor(row, cols, 'Código'), 30)
        titulo = _texto(_valor(row, cols, 'Título do vídeo'), 200)
        nivel = _inteiro(_valor(row, cols, 'Obrigatório a partir do nível'))
        if not familia or not codigo or not titulo or not nivel:
            continue
        conteudos.append({
            'familia': familia, 'codigo': codigo,
            'modulo': _texto(_valor(row, cols, 'Módulo'), 150),
            'titulo': titulo,
            'categoria': _texto(_valor(row, cols, 'Categoria'), 100),
            'nivel_minimo': nivel,
            'objetivo': _texto(_valor(row, cols, 'Objetivo do conteúdo')),
        })

    enquadramentos = []
    for row, cols in _linhas(_aba(wb, 'Enquadramento Atual'), 'Nome'):
        nome = _texto(_valor(row, cols, 'Nome'), 200)
        familia = _texto(_valor(row, cols, 'Família sugerida'), 100)
        if not nome or not familia:
            continue
        decisao = _texto(_valor(row, cols, 'Decisão registrada'), 30)
        enquadramentos.append({
            'nome': nome, 'familia': familia,
            'nivel': _inteiro(_valor(row, cols, 'Nível inicial')) or None,
            'cargo_atual_planilha': _texto(_valor(row, cols, 'Cargo atual'), 150),
            'salario_base_atual': _numero(_valor(row, cols, 'Salário base atual')),
            'complementos_atuais': _numero(_valor(row, cols, 'Complementos/HE/DSR atuais')),
            'total_atual': _numero(_valor(row, cols, 'Total atual')),
            'cargo_proposto': _texto(_valor(row, cols, 'Cargo proposto'), 150),
            'salario_base_alvo': _numero(_valor(row, cols, 'Base alvo mensal')),
            'complemento_funcao_alvo': _numero(_valor(row, cols, 'Complemento função alvo')),
            'total_alvo': _numero(_valor(row, cols, 'Total alvo')),
            'status_cenario': _texto(_valor(row, cols, 'Status do cenário total'), 80),
            'nota_transicao': _texto(_valor(row, cols, 'Nota de transição')),
            'fonte_competencia': _texto(_valor(row, cols, 'Fonte / competência do cenário atual')),
            'decisao': decisao if decisao in DECISOES else None,
        })

    validacoes = []
    for row, cols in _linhas(_aba(wb, 'Guia e Controle'), 'Validação'):
        tema = _texto(_valor(row, cols, 'Validação'), 200)
        ordem = _inteiro(_valor(row, cols, '#'))
        if not ordem or not tema:
            if validacoes:
                break
            continue
        validacoes.append({
            'ordem': ordem, 'tema': tema,
            'motivo': _texto(_valor(row, cols, 'Motivo')),
            'responsavel': _texto(_valor(row, cols, 'Responsável'), 200),
            'estado': _texto(_valor(row, cols, 'Estado'), 50),
            'evidencia_esperada': _texto(_valor(row, cols, 'Evidência esperada')),
            'bloqueia_implantacao': _texto(_valor(row, cols, 'Bloqueia implantação?'), 80),
        })

    referencia = _texto(_aba(wb, 'Guia e Controle')['A2'].value, 100)
    return {'sha256': hashlib.sha256(raw).hexdigest(), 'referencia': referencia,
            'faixas': faixas, 'regras': regras, 'conteudos': conteudos,
            'enquadramentos': enquadramentos, 'validacoes': validacoes}


def _vincular_funcionarios(dados):
    por_nome = defaultdict(list)
    for funcionario in Funcionario.query.order_by(Funcionario.id).all():
        por_nome[_norm(funcionario.nome)].append(funcionario)
    avisos, vinculados = [], 0
    for linha in dados['enquadramentos']:
        candidatos = por_nome.get(_norm(linha['nome']), [])
        ativos = [f for f in candidatos if f.ativo]
        escolhido = ativos[0] if len(ativos) == 1 else (
            candidatos[0] if len(candidatos) == 1 else None)
        linha['funcionario_id'] = escolhido.id if escolhido else None
        if escolhido:
            vinculados += 1
        elif candidatos:
            avisos.append(f'{linha["nome"]}: há mais de uma ficha ativa com este nome.')
        else:
            avisos.append(f'{linha["nome"]}: funcionário não encontrado no RH.')
    return vinculados, avisos


def _vincular_videos(dados):
    por_chave, por_titulo = defaultdict(list), defaultdict(list)
    for video in TreinoVideo.query.all():
        por_chave[(_norm(video.trilha.nome), _norm(video.titulo))].append(video)
        por_titulo[_norm(video.titulo)].append(video)
    ligados = set()
    for linha in dados['conteudos']:
        candidatos = por_chave.get((_norm(linha['modulo']), _norm(linha['titulo'])), [])
        if len(candidatos) != 1:
            candidatos = por_titulo.get(_norm(linha['titulo']), [])
        video = candidatos[0] if len(candidatos) == 1 else None
        linha['treino_video_id'] = video.id if video else None
        if video:
            ligados.add(video.id)
    return len(ligados)


def _prever_vinculos_cargos(dados):
    """Conta vínculos por nome sem alterar a tabela contratual de cargos."""
    existentes = {_norm(c.nome) for c in Cargo.query.all()}
    nomes = {_norm(f['cargo_proposto']) for f in dados['faixas']}
    nomes.discard('')
    return {
        'cargos_vinculados': len(nomes & existentes),
        'cargos_a_criar': len(nomes - existentes),
    }


def cargo_da_faixa(familia, nivel):
    """Devolve o Cargo real ligado a uma família/nível do plano."""
    if not familia or not nivel:
        return None
    return (Cargo.query
            .join(PlanoCarreiraCargoVinculo,
                  PlanoCarreiraCargoVinculo.cargo_id == Cargo.id)
            .join(PlanoCarreiraFaixa,
                  PlanoCarreiraFaixa.id ==
                  PlanoCarreiraCargoVinculo.faixa_id)
            .filter(PlanoCarreiraFaixa.familia == familia,
                    PlanoCarreiraFaixa.nivel == nivel)
            .first())


def aplicar_cargo_aprovado(enquadramento):
    """Aplica no funcionário o Cargo ligado à faixa já aprovada."""
    cargo = cargo_da_faixa(enquadramento.familia, enquadramento.nivel)
    funcionario = enquadramento.funcionario
    if not cargo or not funcionario:
        return None
    funcionario.cargo = cargo
    funcionario.funcao = cargo.nome
    funcionario.salario_base = cargo.salario_base
    return cargo


def sincronizar_enquadramento_com_cargo(funcionario):
    """Mantém o plano coerente quando o cargo real é trocado na ficha."""
    enquadramento = funcionario.enquadramento_carreira
    if not enquadramento or not funcionario.cargo_id:
        return None
    faixa = (PlanoCarreiraFaixa.query
             .join(PlanoCarreiraCargoVinculo,
                   PlanoCarreiraCargoVinculo.faixa_id ==
                   PlanoCarreiraFaixa.id)
             .filter(PlanoCarreiraCargoVinculo.cargo_id ==
                     funcionario.cargo_id)
             .first())
    if not faixa:
        return None
    enquadramento.familia = faixa.familia
    enquadramento.nivel = faixa.nivel
    enquadramento.cargo_proposto = faixa.cargo_proposto
    enquadramento.salario_base_alvo = faixa.equivalente_mensal or 0
    enquadramento.complemento_funcao_alvo = faixa.complemento_funcao or 0
    enquadramento.total_alvo = faixa.total_alvo or 0
    enquadramento.decisao = 'Aprovado'
    return faixa


def prever(raw: bytes):
    dados = ler(raw)
    vinculados, avisos = _vincular_funcionarios(dados)
    videos = _vincular_videos(dados)
    cargos = _prever_vinculos_cargos(dados)
    decisoes_atuais = {e.funcionario_id: e.decisao
                       for e in PlanoCarreiraEnquadramento.query.all()
                       if e.decisao}
    dados['avisos'] = avisos
    dados['resumo'] = {
        'faixas': len(dados['faixas']),
        'familias': len({f['familia'] for f in dados['faixas']}),
        'regras': len(dados['regras']), 'conteudos': len(dados['conteudos']),
        'videos_vinculados': videos,
        'pessoas_planilha': len(dados['enquadramentos']),
        'pessoas_vinculadas': vinculados,
        'validacoes': len(dados['validacoes']),
        **cargos,
        'aprovacoes_a_aplicar': sum(
            1 for e in dados['enquadramentos']
            if e.get('funcionario_id')
            and (e.get('decisao') == 'Aprovado'
                 or (not e.get('decisao') and decisoes_atuais.get(
                     e['funcionario_id']) == 'Aprovado'))
            and e.get('nivel')),
    }
    return dados


def aplicar(raw: bytes, nome_arquivo: str, usuario_id=None):
    dados = prever(raw)
    decisoes_atuais = {e.funcionario_id: e.decisao
                       for e in PlanoCarreiraEnquadramento.query.all()
                       if e.decisao}
    try:
        for model in (PlanoCarreiraConteudo, PlanoCarreiraEnquadramento,
                      PlanoCarreiraCargoVinculo,
                      PlanoCarreiraFaixa, PlanoCarreiraRegra,
                      PlanoCarreiraValidacao):
            model.query.delete(synchronize_session=False)
        PlanoCarreiraImportacao.query.delete(synchronize_session=False)
        db.session.expunge_all()
        imp = PlanoCarreiraImportacao(
            nome_arquivo=_texto(nome_arquivo, 255) or 'plano-carreira.xlsx',
            sha256=dados['sha256'], referencia=dados['referencia'],
            importado_por_id=usuario_id)
        db.session.add(imp)
        db.session.flush()
        # Cada faixa recebe um Cargo real. Cargos que já existem são
        # preservados (inclusive o salário vigente); os ausentes nascem com a
        # base de referência da faixa e passam a aparecer no seletor do RH.
        cargos_por_nome = {_norm(c.nome): c for c in Cargo.query.all()}
        cargos_por_faixa = {}
        for item in dados['faixas']:
            faixa = PlanoCarreiraFaixa(importacao_id=imp.id, **item)
            db.session.add(faixa)
            chave_nome = _norm(item['cargo_proposto'])
            cargo = cargos_por_nome.get(chave_nome)
            if cargo is None:
                cargo = Cargo(
                    nome=item['cargo_proposto'],
                    salario_base=item['equivalente_mensal'] or 0,
                    descricao=f'Faixa {item["familia"]} N{item["nivel"]}',
                    ativo=True,
                )
                db.session.add(cargo)
                cargos_por_nome[chave_nome] = cargo
            db.session.flush()
            db.session.add(PlanoCarreiraCargoVinculo(
                faixa_id=faixa.id, cargo_id=cargo.id))
            cargos_por_faixa[(item['familia'], item['nivel'])] = cargo
        for item in dados['regras']:
            db.session.add(PlanoCarreiraRegra(importacao_id=imp.id, **item))
        for item in dados['conteudos']:
            db.session.add(PlanoCarreiraConteudo(importacao_id=imp.id, **item))
        for item in dados['validacoes']:
            db.session.add(PlanoCarreiraValidacao(importacao_id=imp.id, **item))
        for item in dados['enquadramentos']:
            if not item.get('funcionario_id'):
                continue
            linha = dict(item)
            linha.pop('nome', None)
            if not linha.get('decisao'):
                linha['decisao'] = decisoes_atuais.get(linha['funcionario_id'])
            enquadramento = PlanoCarreiraEnquadramento(
                importacao_id=imp.id, **linha)
            db.session.add(enquadramento)
            # "Aprovado" é a decisão final do dono: neste ponto o cargo real
            # passa a ser o Cargo ligado à faixa. Os demais enquadramentos
            # continuam apenas como proposta.
            if linha.get('decisao') == 'Aprovado' and linha.get('nivel'):
                cargo = cargos_por_faixa.get(
                    (linha['familia'], linha['nivel']))
                funcionario = db.session.get(
                    Funcionario, linha['funcionario_id'])
                if cargo and funcionario:
                    funcionario.cargo = cargo
                    funcionario.funcao = cargo.nome
                    funcionario.salario_base = cargo.salario_base
        db.session.commit()
    except Exception:
        db.session.rollback()
        raise
    return dados['resumo'] | {'avisos': dados['avisos'], 'importacao_id': imp.id}

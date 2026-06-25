"""Geracao de relatorio de pedidos em XLSX e PDF."""

import io
import logging
import math

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.pdf import PadariaPDF

logger = logging.getLogger(__name__)

# Magic bytes dos formatos que o fpdf2 aceita — pra validar que o download
# trouxe uma IMAGEM, nao a pagina HTML de preview do Dropbox.
_IMG_MAGIC = (b'\xff\xd8\xff', b'\x89PNG\r\n\x1a\n', b'GIF87a', b'GIF89a')


def _money(v):
    return f'R$ {v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


def _eh_imagem(conteudo, content_type):
    """True se o conteudo baixado e realmente uma imagem (nao HTML)."""
    if content_type.startswith('image/'):
        return True
    return any(conteudo.startswith(m) for m in _IMG_MAGIC)


def _foto_bytes(foto):
    """Le bytes da foto pra embutir no PDF.

    Ordem de tentativa (cada uma logada em caso de falha):
    1. **API autenticada via storage_path** — `dropbox_storage.baixar(path)`.
       SEMPRE funciona se temos token + path; nao depende de shared link,
       CDN, raw vs preview, ou User-Agent. Eh o canonico.
    2. **HTTP no shared link** (legado/compat) — User-Agent de navegador,
       URL normalizada pra `raw=1`. Valida magic bytes / Content-Type pra
       rejeitar pagina HTML de preview do Dropbox.
    3. **BLOB legado** (`foto.imagem`) — fotos pre-M6.

    Bug 24/06/2026: fix via User-Agent + raw=1 nao bastou em prod (motivo
    nao confirmado — pode ser link de preview que ignora raw, ou CDN com
    rate limit). Solucao canonica = baixar via API autenticada, tirando o
    CDN publico do caminho critico.
    """
    from app.services import dropbox_storage
    from app.services.dropbox_storage import _converter_para_raw

    storage_path = getattr(foto, 'imagem_storage_path', None)
    if storage_path:
        bytes_ = dropbox_storage.baixar(storage_path)
        if bytes_:
            return bytes_
        logger.warning(
            'foto %s: API Dropbox falhou pra %s — tentando shared link',
            getattr(foto, 'id', '?'), storage_path)

    url = foto.imagem_url
    if url:
        try:
            r = requests.get(
                _converter_para_raw(url), timeout=15,
                headers={'User-Agent':
                         'Mozilla/5.0 (compatible; PadariaPDF/1.0)'})
            conteudo = r.content or b''
            ct = (r.headers.get('Content-Type') or '').lower()
            if r.status_code == 200 and _eh_imagem(conteudo, ct):
                return conteudo
            logger.warning(
                'foto %s: shared link nao retornou imagem '
                '(status=%s content_type=%s len=%s) — usando BLOB',
                getattr(foto, 'id', '?'), r.status_code, ct, len(conteudo))
        except Exception:  # noqa: BLE001
            logger.exception('foto %s: erro no shared link pro PDF',
                             getattr(foto, 'id', '?'))
    return foto.imagem  # BLOB legado (pode ser None apos M6)


def _fotos_conferencia(p, etapa=None):
    """Fotos do QR (PedidoItemFoto) do pedido, por SKU.

    Modelo DIFERENTE de `FotoRecebimento` (= `p.fotos`, upload manual pela
    web): essas sao as tiradas no fluxo de conferencia via QR code —
    `saida` (industria->motorista) e `entrega` (motorista->loja). Eram
    invisiveis no PDF ate 25/06/2026 porque o relatorio so olhava `p.fotos`.

    `etapa`: filtra ('saida' | 'entrega'); None = ambas.

    Retorna [(foto, legenda)] ordenado por etapa; legenda = item + etapa.
    """
    out = []
    for item in p.itens:
        for f in (item.fotos_conferencia or []):
            if etapa and f.etapa != etapa:
                continue
            nome = (item.nome_item or '')[:18]
            out.append((f, f'{nome} ({f.etapa or "?"})'))
    out.sort(key=lambda t: (t[0].etapa or '', t[1]))
    return out


def _render_fotos(pdf, fotos, titulo='Fotos do recebimento', legendas=None,
                  largura=45, altura=35, por_linha=4, margem=2):
    """Renderiza miniaturas das fotos em grade dentro do PDF.

    Quebra de pagina automatica quando o bloco nao cabe no restante.
    Fotos podem estar no Dropbox (M6+) ou BLOB legado.

    `legendas`: lista paralela a `fotos` (str por foto). Quando presente,
    escreve a legenda embaixo de cada miniatura — usado nas fotos de
    conferencia por SKU pra identificar item + etapa.
    """
    pdf.set_font('Helvetica', 'I', 8)
    pdf.cell(0, 4, f'{titulo} ({len(fotos)}):', new_x='LMARGIN', new_y='NEXT')

    h_leg = 4 if legendas else 0
    bloco = altura + h_leg
    x0 = pdf.l_margin
    y = pdf.get_y()
    for i, foto in enumerate(fotos):
        col = i % por_linha
        if col == 0 and i > 0:
            y += bloco + margem
        if y + bloco > pdf.h - pdf.b_margin - 5:
            pdf.add_page()
            y = pdf.get_y()
        x = x0 + col * (largura + margem)
        try:
            bytes_ = _foto_bytes(foto)
            if not bytes_:
                raise ValueError('foto sem bytes')
            pdf.image(io.BytesIO(bytes_), x=x, y=y, w=largura, h=altura)
        except Exception:
            pdf.set_xy(x, y)
            pdf.cell(largura, altura, '[foto invalida]', border=1, align='C')
        if legendas:
            pdf.set_xy(x, y + altura)
            pdf.set_font('Helvetica', '', 6)
            pdf.cell(largura, h_leg, legendas[i][:32], align='C')
    pdf.set_y(y + bloco + margem)


def gerar_xlsx_pedidos(loja_nome, de, ate, pedidos, totais, por_item):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Pedidos'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='37474F', end_color='37474F', fill_type='solid')
    bold = Font(bold=True)
    thin = Side(border_style='thin', color='B0B0B0')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    ws['A1'] = f'Relatorio de Pedidos - {loja_nome}'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:H1')
    ws['A2'] = f'Periodo: {de.strftime("%d/%m/%Y")} a {ate.strftime("%d/%m/%Y")}'
    ws.merge_cells('A2:H2')
    ws['A3'] = ('Valores em PRECO INTERNO (transferencia loja->industria) '
                '- nao e balcao/site/atacado. Item sem interno = R$ 0,00.')
    ws['A3'].font = Font(italic=True, size=9, color='666666')
    ws.merge_cells('A3:H3')

    ws['A4'] = 'Pedidos entregues'
    ws['B4'] = totais['qtd_pedidos']
    ws['D4'] = 'Valor total'
    ws['E4'] = _money(totais['valor_total'])
    ws['G4'] = 'Divergencias'
    ws['H4'] = totais['divergencias']
    for c in ('A4', 'D4', 'G4'):
        ws[c].font = bold

    ws.append([])
    ws.append(['Resumo por item'])
    ws[ws.max_row][0].font = bold
    head_row = ['Item', 'Qtd pedida', 'Qtd recebida', 'Valor']
    ws.append(head_row)
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    for nome, d in por_item.items():
        ws.append([nome, d['quantidade'], d['recebido'], _money(d['valor'])])
        for cell in ws[ws.max_row]:
            cell.border = border

    ws.append([])
    ws.append(['Detalhamento por pedido'])
    ws[ws.max_row][0].font = bold
    head_row2 = ['Data', 'Pedido', 'Item', 'Qtd pedida', 'Qtd recebida', 'Preco unit.', 'Subtotal', 'Divergente', 'Fotos']
    ws.append(head_row2)
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for p_info in pedidos:
        p = p_info['p']
        # Total de fotos = recebimento manual (p.fotos) + conferencia QR.
        n_fotos = len(p.fotos) + len(_fotos_conferencia(p))
        for l in p_info['linhas']:
            ws.append([
                p.data_entrega.strftime('%d/%m/%Y') if p.data_entrega else '',
                f'#{p.id}',
                l['nome'],
                l['quantidade'],
                l['recebido'],
                _money(l['preco']),
                _money(l['subtotal']),
                'SIM' if l['divergente'] else '',
                n_fotos if n_fotos else '',
            ])
            for cell in ws[ws.max_row]:
                cell.border = border

    ws.append([])
    ws.append(['', '', '', '', '', 'TOTAL', _money(totais['valor_total'])])
    for cell in ws[ws.max_row]:
        cell.font = bold

    for col_idx, width in enumerate([12, 10, 38, 12, 14, 14, 14, 12, 8], start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Layout do relatorio PDF: manter cada pedido inteiro numa pagina ──────
#
# Bug (25/06/2026, /pedidos/relatorio?formato=pdf, visto na Loja Nebraska):
# a tabela de itens de um pedido rachava entre paginas — o ultimo item e o
# "Subtotal do pedido" caiam orfaos no topo da pagina seguinte. Causa: as
# quebras eram decididas por-LINHA (pdf.get_y() > 275), entao o bloco fluia
# ate o fim da pagina e quebrava onde calhasse. Fix: estimar a altura do
# bloco do pedido (cabecalho + itens + subtotal [+ fotos]) ANTES de desenhar
# e empurrar o pedido inteiro pra proxima pagina quando nao cabe.
#
# Alturas (mm) das celulas desenhadas em _render_pedido — manter em sincronia
# com o desenho abaixo.
_H_HEADER_PEDIDO = 5
_H_LINHA_ITEM = 4
_H_SUBTOTAL = 5
_H_FOLGA_PEDIDO = 2
_H_TITULO_SECAO = 6


def _altura_bloco_pedido(n_linhas):
    """Altura estimada (mm) do bloco de texto de um pedido: cabecalho do
    pedido + uma linha por item + linha de subtotal + folga final."""
    return (_H_HEADER_PEDIDO + n_linhas * _H_LINHA_ITEM
            + _H_SUBTOTAL + _H_FOLGA_PEDIDO)


def _altura_grade_fotos(n, com_legenda):
    """Altura estimada (mm) de uma grade de fotos como _render_fotos a
    desenha: titulo + ceil(n/por_linha) linhas de miniaturas. Espelha a
    geometria de _render_fotos (altura=35, por_linha=4, margem=2) pra a
    decisao de quebra manter as fotos junto do pedido quando cabem."""
    if not n:
        return 0.0
    bloco = 35 + (4 if com_legenda else 0)  # miniatura + legenda opcional
    linhas = math.ceil(n / 4)
    return 4 + linhas * (bloco + 2)  # 4 = titulo; 2 = margem entre linhas


def _render_pedido(pdf, p_info, incluir_fotos, conf, limite, altura_pagina,
                   titulo_secao=None):
    """Desenha o bloco de um pedido (cabecalho + itens + subtotal + fotos),
    decidindo a quebra de pagina ANTES pra nao rachar o pedido entre paginas.

    `titulo_secao`: se setado, imprime esse titulo de secao logo acima do
    pedido e dentro da mesma decisao de quebra (evita o titulo orfao no
    rodape quando o 1o pedido pula de pagina).

    Retorna dict com a pagina de inicio do pedido e a pagina do subtotal —
    usado em teste pra travar que pedidos que cabem numa pagina nao racham
    (pagina_inicio == pagina_subtotal)."""
    p = p_info['p']
    n_fotos = len(p.fotos) + len(conf)

    alt_texto = _altura_bloco_pedido(len(p_info['linhas']))
    if titulo_secao:
        alt_texto += _H_TITULO_SECAO
    alt_fotos = 0.0
    if incluir_fotos:
        if p.fotos:
            alt_fotos += _altura_grade_fotos(len(p.fotos), com_legenda=False)
        if conf:
            alt_fotos += _altura_grade_fotos(len(conf), com_legenda=True)
    alt_total = alt_texto + alt_fotos
    restante = limite - pdf.get_y()

    # Mantem o pedido junto. Cabe tudo aqui -> fica. Cabe numa pagina limpa
    # -> pula a pagina (pedido + fotos juntos). Nem numa pagina inteira cabe
    # (pedido gigante ou muitas fotos) -> garante ao menos o bloco de TEXTO
    # inteiro e deixa as fotos fluirem (elas ja paginam sozinhas).
    if alt_total <= restante:
        pass
    elif alt_total <= altura_pagina:
        pdf.add_page()
    elif alt_texto > restante:
        pdf.add_page()

    if titulo_secao:
        pdf.set_font('Helvetica', 'B', 10)
        pdf.cell(0, _H_TITULO_SECAO, titulo_secao,
                 new_x='LMARGIN', new_y='NEXT')

    pagina_inicio = pdf.page

    pdf.set_fill_color(235, 235, 235)
    pdf.set_font('Helvetica', 'B', 9)
    data_str = p.data_entrega.strftime('%d/%m/%Y') if p.data_entrega else '-'
    div_str = '  [DIVERGENCIA]' if p.tem_divergencia else ''
    fotos_str = (f'  ({n_fotos} foto{"s" if n_fotos != 1 else ""})'
                 if n_fotos else '')
    pdf.cell(0, _H_HEADER_PEDIDO,
             f'Pedido #{p.id}  -  {data_str}{div_str}{fotos_str}',
             fill=True, border=1, new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', '', 8)
    for l in p_info['linhas']:
        pdf.cell(85, _H_LINHA_ITEM, l['nome'][:48], border='LR')
        pdf.cell(20, _H_LINHA_ITEM, f'{l["recebido"]}x', border='LR', align='C')
        pdf.cell(30, _H_LINHA_ITEM, _money(l['preco']), border='LR', align='R')
        pdf.cell(35, _H_LINHA_ITEM, _money(l['subtotal']), border='LR',
                 align='R', new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', 'B', 9)
    pagina_subtotal = pdf.page
    pdf.cell(135, _H_SUBTOTAL, 'Subtotal do pedido', border=1, align='R')
    pdf.cell(35, _H_SUBTOTAL, _money(p_info['subtotal']), border=1, align='R',
             new_x='LMARGIN', new_y='NEXT')

    if incluir_fotos:
        if p.fotos:
            _render_fotos(pdf, list(p.fotos),
                          titulo='Fotos do recebimento (upload manual)')
        if conf:
            _render_fotos(pdf, [f for f, _ in conf],
                          titulo='Fotos da conferencia por item (QR)',
                          legendas=[lg for _, lg in conf])

    pdf.ln(_H_FOLGA_PEDIDO)
    return {'id': p.id, 'pagina_inicio': pagina_inicio,
            'pagina_subtotal': pagina_subtotal,
            'n_linhas': len(p_info['linhas'])}


def montar_pdf_pedidos(loja_nome, de, ate, pedidos, totais, por_item,
                       incluir_fotos=False, etapa_foto=None):
    """Monta o FPDF do relatorio de pedidos. Separado de gerar_ pra teste
    poder inspecionar `pdf.layout_pedidos` (pagina de inicio/subtotal de cada
    pedido) e travar que pedidos que cabem numa pagina nao racham."""
    pdf = PadariaPDF(orientation='P', unit='mm', format='A4')
    pdf.add_page()
    limite = pdf.h - pdf.b_margin        # y maximo util da pagina (~277mm)
    altura_pagina = limite - pdf.get_y()  # area util abaixo do cabecalho fixo

    pdf.set_font('Helvetica', 'B', 12)
    pdf.cell(0, 7, f'Relatorio de Pedidos - {loja_nome}', align='C', new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', '', 9)
    pdf.cell(0, 5, f'Periodo: {de.strftime("%d/%m/%Y")} a {ate.strftime("%d/%m/%Y")}',
             align='C', new_x='LMARGIN', new_y='NEXT')
    pdf.set_font('Helvetica', 'I', 8)
    pdf.cell(0, 4, 'Valores em preco INTERNO (transferencia loja-industria) - '
             'nao e balcao/site/atacado.',
             align='C', new_x='LMARGIN', new_y='NEXT')
    pdf.ln(3)

    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(60, 6, f'Pedidos entregues: {totais["qtd_pedidos"]}', new_x='RIGHT')
    pdf.cell(70, 6, f'Valor total: {_money(totais["valor_total"])}', new_x='RIGHT')
    pdf.cell(60, 6, f'Divergencias: {totais["divergencias"]}', new_x='LMARGIN', new_y='NEXT')
    pdf.ln(3)

    # Resumo por item — tabela continua; quebra normal por linha (nao e um
    # bloco de pedido, entao dividir entre paginas e esperado).
    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(0, 6, 'Resumo por item', new_x='LMARGIN', new_y='NEXT')
    pdf.set_fill_color(55, 71, 79)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Helvetica', 'B', 9)
    pdf.cell(80, 6, 'Item', fill=True, border=1)
    pdf.cell(30, 6, 'Qtd pedida', fill=True, border=1, align='C')
    pdf.cell(30, 6, 'Qtd recebida', fill=True, border=1, align='C')
    pdf.cell(40, 6, 'Valor', fill=True, border=1, align='R', new_x='LMARGIN', new_y='NEXT')
    pdf.set_text_color(0, 0, 0)
    pdf.set_font('Helvetica', '', 9)
    for nome, d in por_item.items():
        if pdf.get_y() > limite - 7:
            pdf.add_page()
        pdf.cell(80, 5, nome[:42], border=1)
        pdf.cell(30, 5, str(d['quantidade']), border=1, align='C')
        pdf.cell(30, 5, str(d['recebido']), border=1, align='C')
        pdf.cell(40, 5, _money(d['valor']), border=1, align='R', new_x='LMARGIN', new_y='NEXT')
    pdf.ln(3)

    # Detalhamento por pedido — cada pedido fica inteiro numa pagina (o titulo
    # da secao viaja junto do 1o pedido pra nao ficar orfao no rodape).
    layout = []
    for i, p_info in enumerate(pedidos):
        p = p_info['p']
        # Fotos de conferencia (QR) so sao coletadas quando vao ser renderizadas
        # — evita N+1 query no relatorio comum (sem fotos).
        conf = _fotos_conferencia(p, etapa_foto) if incluir_fotos else []
        titulo = 'Detalhamento por pedido' if i == 0 else None
        layout.append(_render_pedido(pdf, p_info, incluir_fotos, conf,
                                     limite, altura_pagina, titulo_secao=titulo))

    if not pedidos:
        # Sem pedidos, ainda imprime o titulo da secao (consistencia visual).
        pdf.set_font('Helvetica', 'B', 10)
        pdf.cell(0, 6, 'Detalhamento por pedido', new_x='LMARGIN', new_y='NEXT')

    if pdf.get_y() > limite - 17:
        pdf.add_page()
    pdf.ln(2)
    pdf.set_font('Helvetica', 'B', 11)
    pdf.set_fill_color(235, 235, 235)
    pdf.cell(135, 7, 'TOTAL GERAL', border=1, align='R', fill=True)
    pdf.cell(35, 7, _money(totais['valor_total']), border=1, align='R', fill=True,
             new_x='LMARGIN', new_y='NEXT')

    pdf.layout_pedidos = layout
    return pdf


def gerar_pdf_pedidos(loja_nome, de, ate, pedidos, totais, por_item,
                      incluir_fotos=False, etapa_foto=None):
    pdf = montar_pdf_pedidos(loja_nome, de, ate, pedidos, totais, por_item,
                             incluir_fotos=incluir_fotos, etapa_foto=etapa_foto)
    buf = io.BytesIO()
    pdf.output(buf)
    buf.seek(0)
    return buf

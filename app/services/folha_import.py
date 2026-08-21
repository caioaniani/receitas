"""Importação da FOLHA DE PAGAMENTO (xlsx da contabilidade) pro RH — 03/08/2026.

Pedido do dono ("Preciso atualizar o RH", planilha 06/2026 anexa). A folha é
MENSAL, então em vez de um acerto manual isto virou tela: upload do xlsx em
/rh/folha → PRÉVIA linha a linha (novo / salário ou cargo mudou / igual /
está no sistema mas fora da folha) → o dono marca o que aplicar → aplica.

Regras (salário é dinheiro — peso especial, CLAUDE.md):
- Match por **CPF** (só dígitos; `Funcionario.cpf` é unique — models/rh.py:35).
- NADA é gravado sem passar pela prévia e pelo checkbox do dono. Não existe
  "aplicar tudo" silencioso.
- Funcionário no sistema que NÃO veio na folha NUNCA é desligado sozinho —
  vira uma seção da prévia com checkbox "marcar desligado" POR PESSOA
  (férias/licença também somem de folha; só o dono sabe a diferença).
- A folha MANDA nos campos que ela traz (salário base, cargo, admissão);
  o resto do cadastro (VT/VR, telefone, loja) fica intacto.

Formato esperado (aba "Funcionários", relatório da contabilidade):
cabeçalho na linha 1 com ao menos Nome, CPF, Cargo, Salário Base; Admissão
opcional. Colunas localizadas por NOME no cabeçalho, não por posição — a
contabilidade às vezes muda a ordem.
"""
import io
import logging
import re
from datetime import date, datetime

from app.extensions import db
from app.models import Cargo, Funcionario

logger = logging.getLogger(__name__)

ABA_FUNCIONARIOS = 'Funcionários'


def _so_digitos(cpf):
    return re.sub(r'\D', '', str(cpf or ''))


def _num(v):
    """Número da célula: openpyxl entrega float pra numérico; string
    '2.130,40' (pt-BR) também aparece em export antigo."""
    if v is None or v == '':
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def _data(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or '').strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def ler_folha(conteudo_bytes):
    """Lê a aba de funcionários do xlsx. Devolve (linhas, avisos).

    linha = {nome, cpf (só dígitos), cargo, salario, admissao (date|None)}.
    Linha sem CPF ou sem salário legível vira AVISO visível — nunca some em
    silêncio (é gente e é dinheiro)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(conteudo_bytes), data_only=True)
    ws = None
    for nome_aba in (ABA_FUNCIONARIOS, 'Funcionarios'):
        if nome_aba in wb.sheetnames:
            ws = wb[nome_aba]
            break
    if ws is None:
        # Fallback: primeira aba que tenha coluna CPF no cabeçalho.
        for cand in wb.worksheets:
            cab = [str(c.value or '').strip().lower()
                   for c in next(cand.iter_rows(max_row=1))]
            if any('cpf' in c for c in cab):
                ws = cand
                break
    if ws is None:
        raise ValueError('Não achei a aba de funcionários (procuro uma aba '
                         '"Funcionários" ou qualquer aba com coluna CPF).')

    cab = [str(c.value or '').strip().lower()
           for c in next(ws.iter_rows(max_row=1))]

    def col(*termos):
        for i, c in enumerate(cab):
            if any(t in c for t in termos):
                return i
        return None

    i_nome = col('nome')
    i_cpf = col('cpf')
    i_cargo = col('cargo')
    i_sal = col('salário base', 'salario base', 'salário', 'salario')
    i_adm = col('admiss')
    if i_nome is None or i_cpf is None or i_sal is None:
        raise ValueError('Cabeçalho inesperado: preciso ao menos das colunas '
                         'Nome, CPF e Salário Base.')

    linhas, avisos = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome = str(row[i_nome] or '').strip()
        if not nome:
            continue
        cpf = _so_digitos(row[i_cpf])
        sal = _num(row[i_sal])
        if len(cpf) != 11:
            avisos.append(f'{nome}: CPF ilegível ("{row[i_cpf]}") — '
                          'linha IGNORADA, confira na planilha.')
            continue
        if sal is None:
            avisos.append(f'{nome}: salário ilegível ("{row[i_sal]}") — '
                          'linha IGNORADA, confira na planilha.')
            continue
        linhas.append({
            'nome': nome[:200],
            'cpf': cpf,
            'cargo': (str(row[i_cargo] or '').strip()[:100]
                      if i_cargo is not None else ''),
            'salario': round(sal, 2),
            'admissao': _data(row[i_adm]) if i_adm is not None else None,
        })
    if not linhas:
        raise ValueError('Nenhuma linha de funcionário legível na planilha.')
    return linhas, avisos


def comparar(linhas):
    """Prévia: o que a folha faria com o RH atual. NÃO grava nada.

    Devolve dict com `novos`, `alterados` (com antes/depois campo a campo),
    `iguais` e `fora_da_folha` (ativos no sistema que a folha não traz)."""
    por_cpf = {_so_digitos(f.cpf): f for f in Funcionario.query.all()}
    novos, alterados, iguais = [], [], []
    cpfs_folha = set()
    for ln in linhas:
        cpfs_folha.add(ln['cpf'])
        f = por_cpf.get(ln['cpf'])
        if f is None:
            novos.append(ln)
            continue
        difs = {}
        if abs((f.salario_base or 0) - ln['salario']) >= 0.005:
            difs['salario'] = (f.salario_base or 0, ln['salario'])
        if ln['cargo'] and (f.funcao or '').strip().upper() != ln['cargo'].upper():
            difs['cargo'] = (f.funcao or '—', ln['cargo'])
        if ln['admissao'] and f.data_admissao != ln['admissao']:
            difs['admissao'] = (f.data_admissao, ln['admissao'])
        alvo = {'id': f.id, 'nome': f.nome, 'cpf': ln['cpf'],
                'linha': ln, 'difs': difs, 'inativo': not f.ativo}
        (alterados if difs or not f.ativo else iguais).append(alvo)
    fora = [f for cpf, f in por_cpf.items()
            if f.ativo and cpf not in cpfs_folha]
    return {'novos': novos, 'alterados': alterados, 'iguais': iguais,
            'fora_da_folha': fora}


def aplicar(escolhas):
    """Grava SÓ o que o dono marcou na prévia. Devolve stats.

    escolhas = {'criar': [linha...], 'atualizar': [linha...],
                'desligar': [func_id...]}. Cada item re-resolvido aqui
    contra o banco (a prévia é só tela — a autoridade é esta função)."""
    stats = {'criados': 0, 'atualizados': 0, 'reativados': 0,
             'desligados': 0, 'erros': []}
    from app.services import rh_cargos

    cargos = Cargo.query.all()
    for ln in escolhas.get('criar', []):
        cpf = _so_digitos(ln.get('cpf'))
        sal = _num(ln.get('salario'))
        if len(cpf) != 11 or sal is None:
            stats['erros'].append(f"criar {ln.get('nome')}: dados inválidos")
            continue
        if Funcionario.query.filter_by(cpf=cpf).first():
            stats['erros'].append(f"criar {ln.get('nome')}: CPF já cadastrado")
            continue
        funcionario = Funcionario(
            nome=str(ln.get('nome') or '')[:200], cpf=cpf,
            funcao=str(ln.get('cargo') or '')[:100] or None,
            salario_base=sal, data_admissao=_data(ln.get('admissao')),
            ativo=True)
        rh_cargos.associar_funcionario(funcionario, cargos)
        db.session.add(funcionario)
        stats['criados'] += 1
    for ln in escolhas.get('atualizar', []):
        cpf = _so_digitos(ln.get('cpf'))
        f = Funcionario.query.filter_by(cpf=cpf).first()
        sal = _num(ln.get('salario'))
        if f is None or sal is None:
            stats['erros'].append(f"atualizar {ln.get('nome')}: não achei")
            continue
        f.salario_base = sal
        if ln.get('cargo'):
            f.funcao = str(ln['cargo'])[:100]
            # A folha atualiza também o vínculo estruturado quando o nome é
            # inequívoco; sem correspondência, preserva a decisão humana que
            # já estava na ficha e deixa o novo nome para revisão no RH.
            cargo = rh_cargos.encontrar_cargo(f.funcao, cargos)
            if cargo:
                f.cargo_id = cargo.id
        adm = _data(ln.get('admissao'))
        if adm:
            f.data_admissao = adm
        if not f.ativo:
            # Voltou pra folha = está trabalhando de novo.
            f.ativo = True
            f.data_demissao = None
            stats['reativados'] += 1
        stats['atualizados'] += 1
    for fid in escolhas.get('desligar', []):
        f = db.session.get(Funcionario, int(fid))
        if f is None:
            continue
        f.ativo = False
        if not f.data_demissao:
            from app.utils import hoje
            f.data_demissao = hoje()
        stats['desligados'] += 1
    db.session.commit()
    return stats

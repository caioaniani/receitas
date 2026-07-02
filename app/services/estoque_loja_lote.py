"""Entrada em lote pro estoque de loja (EstoqueLoja).

Espelha estoque_congelados.py mas com semantica de SOMA:
cada item soma na quantidade existente em vez de sobrescrever.

Itens sem match no catalogo (receita/produto/MP/orfao existente) entram
como EstoqueLoja com nome_pendente preenchido — depois o admin vincula
a um item cadastrado via /pedidos/estoque-loja/vincular.

Saida em lote (manual, lojas sem PDV API) usa VendaMapa (canal='lote') pra
lembrar vinculacoes — vincula um nome uma vez, vale pra sempre. Mesmo mapa
unificado do Seru (canal='seru'), so muda o canal.
"""
import re
import unicodedata

from sqlalchemy import func

from app.extensions import db
from app.models import (
    EstoqueLoja,
    MateriaPrima,
    MovEstoqueLoja,
    Produto,
    Receita,
    VendaMapa,
    VendaMapaUso,
)

EXPANSOES = {
    'cro': 'croissant',
    'tra': 'tradicional',
    'trd': 'mini',
    'int': 'integral',
    'gra': 'graos',
}


def _ascii(s):
    if not s:
        return ''
    nf = unicodedata.normalize('NFD', s)
    return ''.join(c for c in nf if unicodedata.category(c) != 'Mn').lower().strip()


def _expandir_abreviacoes(nome_ascii):
    if not nome_ascii:
        return nome_ascii
    tokens = re.split(r'(\s+)', nome_ascii)
    return ''.join(EXPANSOES.get(t, t) if t.strip() else t for t in tokens)


def parsear_linha(linha):
    """Mesmo parser do estoque_congelados — 'Nome: qtd', 'Nome=qtd', etc."""
    linha = (linha or '').strip()
    if not linha or linha.startswith('#'):
        return None
    m = re.split(r'\s*[:=\-—]\s*', linha, maxsplit=1)
    if len(m) == 2 and m[1]:
        nome, qtd_raw = m[0].strip(), m[1].strip()
    else:
        m2 = re.match(r'^(.+?)\s+([\d.,]+)\s*$', linha)
        if not m2:
            return {'linha': linha, 'erro': 'formato'}
        nome, qtd_raw = m2.group(1).strip(), m2.group(2)
    if not nome:
        return {'linha': linha, 'erro': 'sem_nome'}
    qtd_clean = re.sub(r'[^\d]', '', qtd_raw)
    if not qtd_clean:
        return {'linha': linha, 'nome': nome, 'erro': 'sem_quantidade'}
    try:
        qtd = int(qtd_clean)
    except ValueError:
        return {'linha': linha, 'nome': nome, 'erro': 'quantidade_invalida'}
    if qtd <= 0:
        return {'linha': linha, 'nome': nome, 'erro': 'quantidade_invalida'}
    return {'linha': linha, 'nome': nome, 'quantidade': qtd}


def parsear_lista(texto):
    if not texto:
        return []
    out = []
    for linha in texto.splitlines():
        item = parsear_linha(linha)
        if item:
            out.append(item)
    return out


def _apelidos_confirmados():
    """Apelidos globais de lote: VendaMapa canal='lote' confirmado + mapeado.

    Match exato no nome_externo vira atalho na resolucao (entrada e saida)."""
    apelidos = []
    for mp in VendaMapa.query.filter(
        VendaMapa.canal == 'lote',
        VendaMapa.confirmado_em.isnot(None),
        VendaMapa.ignorar.is_(False),
    ).all():
        if mp.receita_id:
            apelidos.append((mp.nome_externo, _ascii(mp.nome_externo),
                              'receita', mp.receita_id,
                              mp.receita.nome if mp.receita else mp.nome_externo))
        elif mp.produto_id:
            apelidos.append((mp.nome_externo, _ascii(mp.nome_externo),
                              'produto', mp.produto_id,
                              mp.produto.nome if mp.produto else mp.nome_externo))
        elif mp.materia_prima_id:
            apelidos.append((mp.nome_externo, _ascii(mp.nome_externo),
                              'mp', mp.materia_prima_id,
                              mp.materia_prima.nome if mp.materia_prima else mp.nome_externo))
    return apelidos


def _carregar_catalogo(loja_id):
    """Catalogo de match: receitas + produtos + materias-primas + orfaos
    daquela loja (EstoqueLoja com nome_pendente) + apelidos globais
    confirmados (VendaMapa canal='lote' mapeado)."""
    receitas = [(r.id, r.nome, _ascii(r.nome)) for r in Receita.query.all()]
    produtos = [(p.id, p.nome, _ascii(p.nome)) for p in Produto.query.all()]
    materias = [(m.id, m.nome, _ascii(m.nome)) for m in MateriaPrima.ativas().all()]
    orfaos = []
    if loja_id:
        orfaos = [
            (ep.id, ep.nome_pendente, _ascii(ep.nome_pendente))
            for ep in EstoqueLoja.query.filter(
                EstoqueLoja.loja_id == loja_id,
                EstoqueLoja.nome_pendente.isnot(None),
                EstoqueLoja.receita_id.is_(None),
                EstoqueLoja.produto_id.is_(None),
                EstoqueLoja.materia_prima_id.is_(None),
            ).all()
            if ep.nome_pendente
        ]
    apelidos = _apelidos_confirmados()
    return receitas, produtos, materias, orfaos, apelidos


def _matches_para(nome, receitas, produtos, materias, orfaos, apelidos=()):
    """Retorna [{tipo, id, nome, match}]. tipo em receita/produto/mp/pendente."""
    alvo = _ascii(nome)
    if not alvo:
        return []
    out, seen = [], set()

    def add(tipo, _id, nome_real, kind):
        key = (tipo, _id)
        if key in seen:
            return
        seen.add(key)
        out.append({'tipo': tipo, 'id': _id, 'nome': nome_real, 'match': kind})

    # 0. apelido global confirmado — match exato no nome digitado vira atalho.
    for ap_digitado, ap_asc, ap_tipo, ap_id, ap_nome in apelidos:
        if ap_asc == alvo:
            add(ap_tipo, ap_id, ap_nome, 'apelido')
    if out:
        return out

    # 1. exato
    for oid, onome, oasc in orfaos:
        if oasc == alvo:
            add('pendente', oid, onome, 'exato')
    for rid, rnome, rasc in receitas:
        if rasc == alvo:
            add('receita', rid, rnome, 'exato')
    for pid, pnome, pasc in produtos:
        if pasc == alvo:
            add('produto', pid, pnome, 'exato')
    for mid, mnome, masc in materias:
        if masc == alvo:
            add('mp', mid, mnome, 'exato')
    if out:
        return out

    # 2. substring
    for oid, onome, oasc in orfaos:
        if alvo in oasc or oasc in alvo:
            add('pendente', oid, onome, 'fuzzy')
    for rid, rnome, rasc in receitas:
        if alvo in rasc or rasc in alvo:
            add('receita', rid, rnome, 'fuzzy')
    for pid, pnome, pasc in produtos:
        if alvo in pasc or pasc in alvo:
            add('produto', pid, pnome, 'fuzzy')
    for mid, mnome, masc in materias:
        if alvo in masc or masc in alvo:
            add('mp', mid, mnome, 'fuzzy')
    if out:
        return out[:10]

    # 3. abreviacoes
    expandido = _expandir_abreviacoes(alvo)
    if expandido != alvo:
        for oid, onome, oasc in orfaos:
            if expandido in oasc or oasc in expandido:
                add('pendente', oid, onome, 'fuzzy')
        for rid, rnome, rasc in receitas:
            if expandido in rasc or rasc in expandido:
                add('receita', rid, rnome, 'fuzzy')
        for pid, pnome, pasc in produtos:
            if expandido in pasc or pasc in expandido:
                add('produto', pid, pnome, 'fuzzy')
        for mid, mnome, masc in materias:
            if expandido in masc or masc in expandido:
                add('mp', mid, mnome, 'fuzzy')

    return out[:10]


def sugerir_para_pendentes(estoques_pendentes):
    """Pra cada EstoqueLoja pendente, retorna {ep_id: melhor_match}.

    Usado em /pedidos/estoque-loja pra pre-selecionar o dropdown da
    vinculacao — admin so confirma com 1 clique em vez de buscar na lista.
    Retorna {} pra item sem nome_pendente. Match pode ser receita/produto/mp.
    """
    if not estoques_pendentes:
        return {}
    receitas = [(r.id, r.nome, _ascii(r.nome)) for r in Receita.query.all()]
    produtos = [(p.id, p.nome, _ascii(p.nome)) for p in Produto.query.all()]
    materias = [(m.id, m.nome, _ascii(m.nome)) for m in MateriaPrima.ativas().all()]
    apelidos = _apelidos_confirmados()
    out = {}
    for ep in estoques_pendentes:
        nome = ep.nome_pendente if hasattr(ep, 'nome_pendente') else None
        if not nome:
            continue
        matches = _matches_para(nome, receitas, produtos, materias, (), apelidos)
        if matches:
            out[ep.id] = matches[0]
    return out


def _filtro_para_resolvido(loja_id, resolvido):
    """Monta o filtro de EstoqueLoja.filter_by() pra um resolvido."""
    filtro = {'loja_id': loja_id}
    if resolvido['tipo'] == 'receita':
        filtro['receita_id'] = resolvido['id']
    elif resolvido['tipo'] == 'produto':
        filtro['produto_id'] = resolvido['id']
    elif resolvido['tipo'] == 'mp':
        filtro['materia_prima_id'] = resolvido['id']
    return filtro


def resolver_lista(linhas_parseadas, loja_id):
    receitas, produtos, materias, orfaos, apelidos = _carregar_catalogo(loja_id)
    enriq = []
    for item in linhas_parseadas:
        if item.get('erro'):
            enriq.append(item)
            continue
        matches = _matches_para(item['nome'], receitas, produtos, materias, orfaos, apelidos)
        resolvido = matches[0] if matches else None
        atual = 0
        if resolvido and loja_id:
            if resolvido['tipo'] == 'pendente':
                ep = EstoqueLoja.query.get(resolvido['id'])
            else:
                ep = EstoqueLoja.query.filter_by(
                    **_filtro_para_resolvido(loja_id, resolvido)).first()
            atual = ep.quantidade if ep else 0
        enriq.append({
            **item,
            'matches': matches,
            'resolvido': resolvido,
            'estoque_atual': atual,
            'novo': atual + item['quantidade'],  # SOMA
        })
    return enriq


def _get_or_create_map(nome_digitado):
    """Acha VendaMapa de lote por nome (case-insensitive) ou cria novo pendente."""
    nome = (nome_digitado or '').strip()
    if not nome:
        return None
    nome_lower = nome.lower()
    mp = VendaMapa.query.filter(
        VendaMapa.canal == 'lote',
        func.lower(VendaMapa.nome_externo) == nome_lower,
    ).first()
    if mp:
        return mp
    mp = VendaMapa(canal='lote', nome_externo=nome)
    db.session.add(mp)
    db.session.flush()
    return mp


def resolver_lista_saida(linhas_parseadas, loja_id):
    """Resolve usando VendaMapa canal='lote' (cria pendentes pra nomes novos).

    Retorna lista enriquecida com 'map_entry' (VendaMapa),
    'estoque_atual', 'novo', 'faltou'.
    """
    enriq = []
    for item in linhas_parseadas:
        if item.get('erro'):
            enriq.append(item)
            continue
        nome = (item.get('nome') or '').strip()
        if not nome:
            enriq.append(item)
            continue
        mp = _get_or_create_map(nome)
        atual = 0
        novo = 0
        faltou = 0
        qtd = item.get('quantidade') or 0
        if mp and mp.estado == 'mapeado' and loja_id:
            filtro = {'loja_id': loja_id}
            if mp.receita_id:
                filtro['receita_id'] = mp.receita_id
            elif mp.produto_id:
                filtro['produto_id'] = mp.produto_id
            elif mp.materia_prima_id:
                filtro['materia_prima_id'] = mp.materia_prima_id
            el = EstoqueLoja.query.filter_by(**filtro).first()
            atual = (el.quantidade if el else 0) or 0
            qtd_efetiva = int((qtd * float(mp.fator_quantidade or 1.0)) + 1e-9)
            real = min(qtd_efetiva, atual)
            novo = atual - real
            faltou = max(0, qtd_efetiva - atual)
        enriq.append({
            **item,
            'map_entry': mp,
            'estoque_atual': atual,
            'novo': novo,
            'faltou': faltou,
        })
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
    return enriq


def aplicar_saida_lote(itens_resolvidos, loja_id, user, referencia=None):
    """Aplica saida usando VendaMapa canal='lote'. Pendentes/ignorados pulados.

    Mapeados: subtrai qtd*fator do EstoqueLoja correspondente. Se estoque
    insuficiente, baixa ate 0 e registra 'venda_loja_sem_estoque' pra falta.
    """
    ref = (referencia or 'Saida em lote').strip()
    aplicados = []
    ignorados = []

    if not loja_id:
        return {'aplicados': [], 'ignorados': [{'linha': '*', 'motivo': 'sem_loja'}]}

    for item in itens_resolvidos:
        if item.get('erro'):
            ignorados.append({'linha': item.get('linha', '?'), 'motivo': item['erro']})
            continue
        try:
            qtd_sub = int(item['quantidade'])
        except (KeyError, TypeError, ValueError):
            ignorados.append({'linha': item.get('linha', '?'), 'motivo': 'quantidade_invalida'})
            continue
        if qtd_sub <= 0:
            ignorados.append({'linha': item.get('linha', '?'), 'motivo': 'quantidade_invalida'})
            continue
        mp = item.get('map_entry')
        if not mp:
            ignorados.append({'linha': item.get('linha', '?'), 'motivo': 'sem_map'})
            continue
        if mp.ignorar:
            ignorados.append({'linha': item.get('linha', '?'),
                              'nome': mp.nome_externo, 'motivo': 'ignorado'})
            continue
        if mp.estado == 'pendente':
            ignorados.append({'linha': item.get('linha', '?'),
                              'nome': mp.nome_externo, 'motivo': 'pendente'})
            continue

        # Marcador loja<->mapa: a tela de mapeamentos lista as lojas que usaram
        # cada apelido (join VendaMapaUso x Loja). A FRACAO mora no DebitoEstoque
        # do motor unico; este marcador eh so o registro de uso pra UI/auditoria.
        if not VendaMapaUso.query.filter_by(
                venda_mapa_id=mp.id, loja_id=loja_id).first():
            db.session.add(VendaMapaUso(venda_mapa_id=mp.id, loja_id=loja_id))

        # Baixa pelo MOTOR UNICO (baixa_venda): composicao (cesta->componentes;
        # simples->ele mesmo), fator, acumulador de fracao por item fisico.
        from app.services.baixa_venda import aplicar_venda
        res = aplicar_venda(
            loja_id, receita_id=mp.receita_id, produto_id=mp.produto_id,
            materia_prima_id=mp.materia_prima_id, qtd=qtd_sub,
            fator=mp.fator_quantidade, canal='lote',
            referencia=f'{ref} [{mp.nome_externo}]',
            pedido_ref=f'lote:{loja_id}:{mp.id}',
            usuario_id=getattr(user, 'id', None), nome_venda=mp.nome_externo)

        if res['sem_alvo']:
            ignorados.append({'linha': item.get('linha', '?'),
                              'nome': mp.nome_externo, 'motivo': 'sem_alvo'})
            continue
        if res['baixado'] == 0 and res['faltou'] == 0 and res['acumulado']:
            ignorados.append({'linha': item.get('linha', '?'),
                              'nome': mp.nome_externo,
                              'motivo': 'fracao_acumulando'})
            continue
        aplicados.append({
            'nome': mp.nome_externo, 'alvo': mp.alvo_nome, 'tipo': mp.alvo_tipo,
            'baixado': res['baixado'], 'faltou': res['faltou'],
        })

    if aplicados:
        db.session.commit()

    return {'aplicados': aplicados, 'ignorados': ignorados}


def aplicar_entrada_lote(itens_resolvidos, loja_id, user, referencia=None):
    """SOMA a quantidade de cada item ao EstoqueLoja correspondente.

    Itens sem match viram EstoqueLoja pendente. Cria movimento de auditoria
    tipo='entrada_lote' com a quantidade SOMADA.
    """
    ref = (referencia or 'Entrada em lote').strip()
    aplicados = []
    ignorados = []

    if not loja_id:
        return {'aplicados': [], 'ignorados': [{'linha': '*', 'motivo': 'sem_loja'}]}

    for item in itens_resolvidos:
        if item.get('erro'):
            ignorados.append({'linha': item.get('linha', '?'), 'motivo': item['erro']})
            continue
        try:
            qtd_add = int(item['quantidade'])
        except (KeyError, TypeError, ValueError):
            ignorados.append({'linha': item.get('linha', '?'), 'motivo': 'quantidade_invalida'})
            continue
        if qtd_add <= 0:
            ignorados.append({'linha': item.get('linha', '?'), 'motivo': 'quantidade_invalida'})
            continue

        resolvido = item.get('resolvido')
        ep = None

        if resolvido and resolvido.get('id'):
            if resolvido['tipo'] == 'pendente':
                ep = EstoqueLoja.query.get(resolvido['id'])
            else:
                filtro = _filtro_para_resolvido(loja_id, resolvido)
                ep = EstoqueLoja.query.filter_by(**filtro).first()
                if not ep:
                    ep = EstoqueLoja(**filtro, quantidade=0)
                    db.session.add(ep)
                    db.session.flush()
            tipo_resultado = resolvido['tipo']
            nome_resultado = resolvido['nome']
        else:
            nome_digitado = (item.get('nome') or '').strip()
            if not nome_digitado:
                ignorados.append({'linha': item.get('linha', '?'), 'motivo': 'sem_nome'})
                continue
            # Reusa linha pendente ja existente com o mesmo nome — evita
            # fragmentar o estoque em multiplas linhas com mesmo `nome_pendente`.
            ep = EstoqueLoja.query.filter_by(
                loja_id=loja_id, nome_pendente=nome_digitado,
                receita_id=None, produto_id=None, materia_prima_id=None,
            ).first()
            if not ep:
                ep = EstoqueLoja(loja_id=loja_id, nome_pendente=nome_digitado, quantidade=0)
                db.session.add(ep)
                db.session.flush()
            tipo_resultado = 'pendente'
            nome_resultado = nome_digitado

        anterior = ep.quantidade or 0
        ep.quantidade = anterior + qtd_add

        db.session.add(MovEstoqueLoja(
            estoque_loja_id=ep.id,
            tipo='entrada_lote',
            quantidade=qtd_add,
            referencia=f'{ref} (era {anterior}, ficou {ep.quantidade})',
            usuario_id=getattr(user, 'id', None),
        ))

        aplicados.append({
            'nome': nome_resultado,
            'tipo': tipo_resultado,
            'anterior': anterior,
            'novo': ep.quantidade,
            'delta': qtd_add,
        })

    if aplicados:
        db.session.commit()

    return {'aplicados': aplicados, 'ignorados': ignorados}


def gerar_xlsx_template_balanco():
    """Planilha de CONTAGEM em branco (bytes .xlsx) com TODOS os itens que a
    loja pede, pra alguem preencher a quantidade fisica (o "caminho ao
    contrario": em vez de mapear uma folha ja preenchida, a gente entrega a
    folha certa pra preencher).

    Uma aba unica: Categoria | Item | Quantidade (em branco). Lista as receitas
    com `sugerir_pedido_loja` (as que a loja PEDE) + os produtos ativos (granola
    etc., que a loja tambem estoca) numa secao "Produtos". Ordenado por
    categoria e nome. Cabecalho com campos Loja/Data/Responsavel pra preencher a
    mao. Os nomes saem EXATAMENTE como no catalogo — assim a reimportacao casa
    100% (sem depender do fuzzy)."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    receitas = (Receita.query
                .filter(Receita.arquivada_em.is_(None),
                        Receita.sugerir_pedido_loja.isnot(False))
                .all())
    produtos = Produto.query.filter_by(ativo=True).all()

    # (categoria, nome) — categoria vazia da receita cai em "Outros"; produtos
    # ficam todos sob "Produtos" pra separar do que e receita.
    linhas = [((r.categoria or 'Outros').strip() or 'Outros', r.nome)
              for r in receitas]
    linhas += [('Produtos', p.nome) for p in produtos]
    linhas.sort(key=lambda x: (x[0].lower(), (x[1] or '').lower()))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Contagem'

    titulo_font = Font(bold=True, size=14)
    sub_font = Font(italic=True, size=10, color='666666')
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='37474F', end_color='37474F',
                              fill_type='solid')

    ws['A1'] = 'CONTAGEM DE ESTOQUE — O Pão'
    ws['A1'].font = titulo_font
    ws['A2'] = 'Loja: ______________     Data: ____/____/______     Responsável: ______________'
    ws['A2'].font = sub_font
    ws['A3'] = ('Preencha a coluna Quantidade com a contagem física. Não altere '
                'os nomes — eles precisam bater com o catálogo.')
    ws['A3'].font = sub_font

    hrow = 5
    for c, nome in enumerate(('Categoria', 'Item', 'Quantidade'), start=1):
        cell = ws.cell(row=hrow, column=c, value=nome)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    r = hrow + 1
    for categoria, nome in linhas:
        ws.cell(row=r, column=1, value=categoria)
        ws.cell(row=r, column=2, value=nome)
        ws.cell(row=r, column=3, value=None)          # Quantidade em branco
        r += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 44
    ws.column_dimensions['C'].width = 14
    ws.freeze_panes = 'A6'                             # trava cabecalho ao rolar

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def parsear_linha_conferencia(linha):
    """Parser da CONFERENCIA (balanço). Diferente do de entrada:
    - aceita quantidade 0 (regra do dono: '0 = zerar o item');
    - quantidade em branco/ausente NAO e erro — marca 'em_branco' (regra
      'vazio = deixa como está', o item e PULADO, nao mexe);
    - quantidade com unidade/letra (ex '100 g', '2 receitas') vira 'unidade'
      explicito — NAO chuta numero (seria ajuste de estoque errado).
    Separador ':', '=', TAB (colar do Excel) ou 2+ espacos."""
    linha = (linha or '').strip()
    if not linha or linha.startswith('#'):
        return None
    m = re.split(r'\s*[:=\t]\s*|\s{2,}', linha, maxsplit=1)
    if len(m) == 2:
        nome, qtd_raw = m[0].strip(), m[1].strip()
    else:
        m2 = re.match(r'^(.+?)\s+(\S+)\s*$', linha)
        if not m2:
            return {'linha': linha, 'nome': linha, 'erro': 'em_branco'}
        nome, qtd_raw = m2.group(1).strip(), m2.group(2).strip()
    if not nome:
        return {'linha': linha, 'erro': 'sem_nome'}
    if not qtd_raw:
        return {'linha': linha, 'nome': nome, 'erro': 'em_branco'}
    limpo = qtd_raw.replace('.', '').replace(',', '')
    if not re.fullmatch(r'\d+', limpo):
        motivo = 'unidade' if re.search(r'[a-zA-Z]', qtd_raw) else 'quantidade_invalida'
        return {'linha': linha, 'nome': nome, 'erro': motivo}
    return {'linha': linha, 'nome': nome, 'quantidade': int(limpo)}


def parsear_conferencia(texto):
    if not texto:
        return []
    out = []
    for linha in texto.splitlines():
        item = parsear_linha_conferencia(linha)
        if item:
            out.append(item)
    return out


def resolver_conferencia(linhas_parseadas, loja_id):
    """Como resolver_lista, mas semantica de CONFERENCIA (SET): novo = contado,
    diff = contado - atual. Mantem linhas com erro (em_branco/unidade) no
    preview pra o usuario ver o que NAO vai mexer. Read-only."""
    receitas, produtos, materias, orfaos, apelidos = _carregar_catalogo(loja_id)
    enriq = []
    for item in linhas_parseadas:
        if item.get('erro'):
            enriq.append(item)
            continue
        matches = _matches_para(item['nome'], receitas, produtos, materias,
                                orfaos, apelidos)
        resolvido = matches[0] if matches else None
        atual, existe = 0, False
        if resolvido and loja_id:
            if resolvido['tipo'] == 'pendente':
                ep = EstoqueLoja.query.get(resolvido['id'])
            else:
                ep = EstoqueLoja.query.filter_by(
                    **_filtro_para_resolvido(loja_id, resolvido)).first()
            if ep:
                existe, atual = True, (ep.quantidade or 0)
        contado = item['quantidade']
        enriq.append({
            **item, 'matches': matches, 'resolvido': resolvido,
            'estoque_atual': atual, 'existe': existe,
            'novo': contado, 'diff': contado - atual,     # SET, nao soma
        })
    return enriq


def aplicar_conferencia(itens_resolvidos, loja_id, user, referencia=None):
    """SETA cada item ao valor CONTADO (balanço). Registra ajuste_conferencia
    com a diferenca (sistema -> real) — mesma semantica da conferencia da tela.

    Regras do dono:
    - qtd > 0  -> SETA (cria a linha se o item casou mas ainda nao tinha estoque);
    - qtd == 0 -> ZERA o item (registra a baixa);
    - em branco/ausente -> NAO mexe (o parser marca 'em_branco'; ignorado aqui);
    - nome sem match com qtd>0 -> vira EstoqueLoja pendente pra vincular depois;
    - nome sem match (ou item novo) com qtd==0 -> no-op (nada a zerar).
    Item que NAO veio na lista fica intacto (nao iteramos sobre ele)."""
    ref = (referencia or 'Conferência em lote').strip()
    aplicados, ignorados = [], []
    if not loja_id:
        return {'aplicados': [], 'ignorados': [{'linha': '*', 'motivo': 'sem_loja'}]}

    for item in itens_resolvidos:
        if item.get('erro'):
            ignorados.append({'linha': item.get('linha', '?'),
                              'nome': item.get('nome'), 'motivo': item['erro']})
            continue
        try:
            contado = int(item['quantidade'])
        except (KeyError, TypeError, ValueError):
            ignorados.append({'linha': item.get('linha', '?'),
                              'motivo': 'quantidade_invalida'})
            continue
        if contado < 0:
            ignorados.append({'linha': item.get('linha', '?'),
                              'motivo': 'quantidade_invalida'})
            continue

        resolvido = item.get('resolvido')
        if resolvido and resolvido.get('id'):
            if resolvido['tipo'] == 'pendente':
                ep = EstoqueLoja.query.get(resolvido['id'])
            else:
                filtro = _filtro_para_resolvido(loja_id, resolvido)
                ep = EstoqueLoja.query.filter_by(**filtro).first()
                if not ep:
                    if contado == 0:
                        ignorados.append({'linha': item.get('linha', '?'),
                                          'nome': resolvido['nome'],
                                          'motivo': 'zero_sem_estoque'})
                        continue
                    ep = EstoqueLoja(**filtro, quantidade=0)
                    db.session.add(ep)
                    db.session.flush()
            nome_resultado = resolvido['nome']
        else:
            nome_digitado = (item.get('nome') or '').strip()
            if not nome_digitado:
                ignorados.append({'linha': item.get('linha', '?'), 'motivo': 'sem_nome'})
                continue
            if contado == 0:
                ignorados.append({'linha': item.get('linha', '?'),
                                  'nome': nome_digitado, 'motivo': 'zero_sem_estoque'})
                continue
            ep = EstoqueLoja.query.filter_by(
                loja_id=loja_id, nome_pendente=nome_digitado,
                receita_id=None, produto_id=None, materia_prima_id=None,
            ).first()
            if not ep:
                ep = EstoqueLoja(loja_id=loja_id, nome_pendente=nome_digitado,
                                 quantidade=0)
                db.session.add(ep)
                db.session.flush()
            nome_resultado = nome_digitado

        anterior = ep.quantidade or 0
        diff = contado - anterior
        if diff != 0:
            ep.quantidade = contado
            db.session.add(MovEstoqueLoja(
                estoque_loja_id=ep.id, tipo='ajuste_conferencia', quantidade=diff,
                referencia=f'{ref}: sistema {anterior} → real {contado} (diff {diff:+d})',
                usuario_id=getattr(user, 'id', None),
            ))
        aplicados.append({'nome': nome_resultado, 'anterior': anterior,
                          'novo': contado, 'diff': diff})

    if any(a['diff'] != 0 for a in aplicados):
        db.session.commit()
    return {'aplicados': aplicados, 'ignorados': ignorados}

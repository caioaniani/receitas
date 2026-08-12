"""Importa CONTATOS (e-mail + celular) dos funcionários por planilha.

Caso que criou a feature (05/08/2026): a rodada de assinatura eletrônica do
Regulamento Interno precisa do e-mail e do WhatsApp de cada funcionário, e o
canal que sustenta a prova em juízo é o que consta DA FICHA do RH — não uma
lista avulsa. O gerente coletou os contatos numa planilha; esta tela põe a
planilha na ficha, com prévia e marcação item a item (mesmo desenho do
import da folha: a prévia é tela, a autoridade é o `aplicar`).

Regras de peso:
- match é por NOME normalizado (a planilha não tem CPF); homônimo no quadro
  vira AVISO e fica de fora — nunca chutar em quem grava contato;
- campo ilegível (e-mail torto, telefone fixo) vira AVISO e o campo fica
  vazio — o outro campo da linha ainda pode ser aplicado;
- aplicar NUNCA apaga valor existente com vazio;
- nome que não existe no quadro vira PRÉ-CADASTRO pendente (reuso do fluxo
  do QR — o RH promove com o CPF depois); nada cria Funcionario direto;
- linha marcada "desligado" na observação NÃO desliga sozinha: vira
  checkbox explícito na prévia (idioma do import da folha).
"""
import io
import logging
import re
import unicodedata

from app.extensions import db
from app.models import Funcionario

# MESMO validador de e-mail do pré-cadastro DE PROPÓSITO (achado de
# revisão): o `aplicar` delega os novos ao `precadastro.validar`, então uma
# regex local divergente faria a prévia aceitar o que o aplicar recusa.
from app.services.precadastro import _RE_EMAIL

logger = logging.getLogger(__name__)


def _norm_nome(s):
    s = unicodedata.normalize('NFKD', s or '')
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return ' '.join(s.upper().split())


def _telefone_celular(bruto):
    """Só dígitos, sem o 55 e sem o 0 de operadora; '' quando não é celular
    BR plausível."""
    from app.services.wifi_portal import _whatsapp_valido
    tel = re.sub(r'\D', '', str(bruto or ''))
    if tel.startswith('55') and len(tel) in (12, 13):
        tel = tel[2:]
    if tel.startswith('0') and len(tel) == 12:
        # "(011) 98888-7777" — escrita comum com o zero de operadora.
        tel = tel[1:]
    return tel if tel and _whatsapp_valido(tel) else ''


def ler_planilha(conteudo_bytes):
    """Lê o xlsx de contatos. Devolve (linhas, avisos).

    linha = {nome, email, telefone, desligado(bool)}. O cabeçalho é
    PROCURADO nas primeiras linhas (a planilha que geramos pro gerente tem
    legenda na linha 1 e cabeçalho na 3) e as colunas são achadas pelo NOME
    — a ordem não importa. Linha "EXEMPLO" é pulada.
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(conteudo_bytes), data_only=True)
    ws, cab, linha_cab = None, None, None
    for cand in wb.worksheets:
        for i, row in enumerate(cand.iter_rows(min_row=1, max_row=10,
                                               values_only=True), 1):
            cels = [str(c or '').strip().lower() for c in row]
            # Cabeçalho de verdade tem nome e e-mail em CÉLULAS DIFERENTES.
            # A legenda da planilha é uma célula mesclada cujo texto contém
            # as duas palavras — sem esta exigência ela era aceita como
            # cabeçalho e o parser lia a coluna A inteira (bug real, pego
            # validando contra a planilha do gerente).
            i_mail = next((j for j, c in enumerate(cels) if 'mail' in c), None)
            i_nom = next((j for j, c in enumerate(cels)
                          if j != i_mail and ('funcion' in c
                                              or c.startswith('nome'))), None)
            if i_mail is not None and i_nom is not None:
                ws, cab, linha_cab = cand, cels, i
                break
        if ws is not None:
            break
    if ws is None:
        raise ValueError('Não achei o cabeçalho — preciso de colunas de '
                         'nome/funcionário e e-mail em alguma aba.')

    def col(*termos):
        for i, c in enumerate(cab):
            if any(t in c for t in termos):
                return i
        return None

    # Reusa os índices DA DETECÇÃO (achado de revisão): re-procurar aqui
    # podia divergir — num cabeçalho "E-mail do funcionário | Nome", o
    # col('funcion') cairia na coluna do e-mail.
    i_nome, i_email = i_nom, i_mail
    i_tel = col('celular', 'telefone', 'whats')
    i_obs = col('observa')

    def cel(row, i):
        return '' if i is None or i >= len(row) else str(row[i] or '').strip()

    linhas, avisos = [], []
    for row in ws.iter_rows(min_row=linha_cab + 1, values_only=True):
        nome = cel(row, i_nome)
        if not nome or nome.upper().startswith('EXEMPLO'):
            continue
        email = cel(row, i_email).lower()
        obs = cel(row, i_obs)
        desligado = 'desligad' in obs.lower()
        if email and not _RE_EMAIL.match(email):
            if not desligado:      # desligado não assina — aviso seria ruído
                avisos.append(f'{nome}: e-mail ilegível ("{email}") — campo '
                              'IGNORADO, confira na planilha.')
            email = ''
        tel_bruto = cel(row, i_tel)
        telefone = _telefone_celular(tel_bruto)
        if tel_bruto and not telefone and not desligado:
            avisos.append(f'{nome}: celular ilegível ("{tel_bruto}") — não é '
                          'um celular BR válido (DDD + 9 dígitos). Campo '
                          'IGNORADO, confira na planilha.')
        linhas.append({'nome': nome[:200], 'email': email[:150],
                       'telefone': telefone, 'desligado': desligado})
    if not linhas:
        raise ValueError('Nenhuma linha de funcionário legível na planilha.')
    return linhas, avisos


def comparar(linhas):
    """Prévia: o que a planilha faria com as fichas. NÃO grava nada."""
    por_nome = {}
    for f in Funcionario.query.all():
        por_nome.setdefault(_norm_nome(f.nome), []).append(f)

    out = {'atualizar': [], 'iguais': [], 'novos': [], 'desligar': [],
           'avisos': []}
    for ln in linhas:
        candidatos = por_nome.get(_norm_nome(ln['nome']), [])
        if len(candidatos) > 1:
            out['avisos'].append(
                f"{ln['nome']}: há {len(candidatos)} fichas com esse nome — "
                'linha FORA da prévia; acerte direto na ficha.')
            continue
        f = candidatos[0] if candidatos else None
        if f is None:
            if ln['desligado']:
                out['avisos'].append(
                    f"{ln['nome']}: marcado como desligado mas não está no "
                    'quadro — nada a fazer.')
            elif len(ln['nome'].split()) < 2 or not ln['email'] \
                    or not ln['telefone']:
                # A prévia só oferece o que o `aplicar` consegue criar: o
                # pré-cadastro exige nome+sobrenome, e-mail e celular
                # (achado de revisão — antes prometia e recusava depois).
                out['avisos'].append(
                    f"{ln['nome']}: não está no quadro e falta "
                    'nome completo, e-mail ou celular válidos pro '
                    'pré-cadastro — complete na planilha.')
            else:
                out['novos'].append(ln)
            continue
        if ln['desligado']:
            # Limitação aceita: linha desligada com contato preenchido não
            # gera item de atualização — se o desligamento for desmarcado,
            # o contato dela não entra (re-importar sem a marca resolve).
            if f.ativo:
                out['desligar'].append({'id': f.id, 'nome': f.nome})
            else:
                out['iguais'].append({'nome': f.nome,
                                      'motivo': 'já desligado na ficha'})
            continue
        difs = {}
        if ln['email'] and (f.email or '').strip().lower() != ln['email']:
            difs['email'] = ((f.email or '—').strip(), ln['email'])
        # Normaliza a ponta da FICHA com a mesma régua da planilha — ficha
        # gravada com o 55 na frente não pode virar falso "mudou".
        tel_ficha = _telefone_celular(f.telefone) \
            or re.sub(r'\D', '', f.telefone or '')
        if ln['telefone'] and tel_ficha != ln['telefone']:
            difs['telefone'] = (f.telefone or '—', ln['telefone'])
        # `inativo` vai pra tela (achado de revisão): atualizar contato de
        # desligado é permitido (ex-funcionário também assina), mas o dono
        # precisa VER que a ficha está desligada.
        alvo = {'id': f.id, 'nome': f.nome, 'linha': ln, 'difs': difs,
                'inativo': not f.ativo}
        (out['atualizar'] if difs else out['iguais']).append(
            alvo if difs else {'nome': f.nome, 'motivo': 'contatos já batem'})
    return out


def aplicar(escolhas):
    """Grava SÓ o que o dono marcou. Cada item re-validado contra o banco.

    escolhas = {'atualizar': [{'id', 'email', 'telefone'}...],
                'precadastro': [{'nome', 'email', 'telefone'}...],
                'desligar': [func_id...]}
    """
    from app.services import precadastro
    from app.utils import hoje

    stats = {'atualizados': 0, 'precadastros': 0, 'desligados': 0,
             'erros': []}
    for item in escolhas.get('atualizar', []):
        if not isinstance(item, dict):
            # POST forjado com JSON que não é objeto ('"texto"', '[1]') —
            # sem o guard virava 500 (achado de revisão).
            stats['erros'].append('atualizar: linha ilegível, pulada')
            continue
        try:
            f = db.session.get(Funcionario, int(item.get('id')))
        except (TypeError, ValueError):
            f = None
        if f is None:
            stats['erros'].append(f"atualizar {item.get('nome') or item.get('id')}: "
                                  'ficha não encontrada')
            continue
        email = (item.get('email') or '').strip().lower()
        tel = _telefone_celular(item.get('telefone'))
        mudou = False
        if email and _RE_EMAIL.match(email):
            f.email = email[:150]
            mudou = True
        if tel:
            f.telefone = tel[:30]
            mudou = True
        if mudou:
            stats['atualizados'] += 1
        else:
            stats['erros'].append(f'{f.nome}: nenhum contato válido pra gravar')
    for fid in escolhas.get('desligar', []):
        try:
            f = db.session.get(Funcionario, int(fid))
        except (TypeError, ValueError):
            f = None
        if f is None or not f.ativo:
            continue
        f.ativo = False
        if not f.data_demissao:
            f.data_demissao = hoje()
        stats['desligados'] += 1
    # Commit das FICHAS antes dos pré-cadastros (achado de revisão): o
    # `precadastro.criar` commita (e a poda interna dele pode dar rollback)
    # — rodá-lo com mutações de ficha ainda pendentes podia descartá-las em
    # silêncio enquanto o flash reportava sucesso.
    db.session.commit()
    for item in escolhas.get('precadastro', []):
        if not isinstance(item, dict):
            stats['erros'].append('pré-cadastro: linha ilegível, pulada')
            continue
        partes = (item.get('nome') or '').split()
        dados, erro = precadastro.validar(
            partes[0] if partes else '', ' '.join(partes[1:]),
            item.get('email'), item.get('telefone'))
        if erro:
            stats['erros'].append(f"pré-cadastro {item.get('nome')}: {erro}")
            continue
        precadastro.criar(dados)
        stats['precadastros'] += 1
    return stats

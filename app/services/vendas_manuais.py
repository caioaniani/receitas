"""Vendas manuais de loja (sem API PDV) + sugestao de pedido.

Cenario: loja sem integracao automatica (ex: Anesio). Admin pode:
- Colar texto: lista 'Nome: qtd' aplicada numa data unica
- Upload de planilha Excel: linhas (data, produto, quantidade) cobrindo
  qualquer periodo, ate um mes inteiro de uma vez.

Em ambos os casos, cria VendaManualLoja sem mexer no estoque. So historico
pra alimentar previsao + sugestao.

Pra sugerir pedido, junta:
- Vendas reais via VNDA (puxa direto da API por periodo — retroativo)
- Vendas reais via Seru (MovEstoqueLoja tipo='venda_seru*')
- Vendas manuais (VendaManualLoja)

Calcula media diaria no periodo + olha estoque atual + sugere qtd.
"""
import io
import math
from collections import defaultdict
from datetime import date, datetime, time, timedelta

from app.constants import VENDA_TIPOS_LOJA
from app.extensions import db
from app.models import (
    EstoqueLoja,
    MateriaPrima,
    MovEstoqueLoja,
    Produto,
    Receita,
    VendaManualLoja,
)
from app.services import estoque_loja_lote as svc_lote
from app.utils import hoje

VENDAS_REAIS = VENDA_TIPOS_LOJA

# Stats da ultima chamada VNDA (por chave loja_id) — usado pra confirmar
# no painel TV se VNDA realmente foi consultada. Cada chamada de
# `_agregar_vendas_vnda_api` sobrescreve a entrada do caller atual.
ULTIMA_CONSULTA_VNDA = {}


def _agregar_vendas_vnda_api(data_inicio, data_fim, _loja_id=None):
    """Puxa direto da API do VNDA pedidos com data_entrega no intervalo
    e agrega por (tipo, id) via VndaProdutoMap. NAO depende de sync
    previo do cron — funciona retroativo.

    Desempacota cestas: se VndaProdutoMap aponta pra Produto cesta (com
    ProdutoItens), explode em componentes (mesma logica do vnda_sync).
    Ex: 34 Family Box (1 unid) × 5 Croissant + 3 Pao Frances + ... vira
    34*5 croissants + 34*3 pao frances + ... contabilizado.

    Atualiza `ULTIMA_CONSULTA_VNDA[_loja_id or 0]` com stats da chamada
    pra debugar no painel sem mexer no contrato de retorno.

    Retorna (vendas_dict, aviso). vendas_dict = {(tipo, id): qtd_total}.
    """
    from app.models import VndaProdutoMap
    from app.services import vnda as vnda_api
    from app.services.vnda_sync import _componentes_de_cesta

    stats_key = _loja_id or 0
    ULTIMA_CONSULTA_VNDA[stats_key] = {
        'tentou_em': datetime.now(),
        'n_orders_recebidos': 0,
        'n_processados': 0,
        'erro': None,
    }

    try:
        todos = vnda_api._buscar_pedidos_janela(data_inicio, data_fim)
    except vnda_api.VndaUnavailableError as e:
        ULTIMA_CONSULTA_VNDA[stats_key]['erro'] = f'VNDA indisponivel: {e}'
        return {}, f'VNDA indisponivel: {e}'
    except Exception as e:  # noqa: BLE001
        ULTIMA_CONSULTA_VNDA[stats_key]['erro'] = f'VNDA falhou: {type(e).__name__}: {str(e)[:200]}'
        return {}, f'VNDA falhou: {type(e).__name__}: {str(e)[:200]}'

    ULTIMA_CONSULTA_VNDA[stats_key]['n_orders_recebidos'] = len(todos or [])

    STATUS_OK_IGNORAR = {'canceled', 'cancelled', 'cancelado'}
    vendas = defaultdict(int)
    n_processados = 0
    for order in todos or []:
        if not isinstance(order, dict):
            continue
        de = vnda_api._extrair_data_entrega(order)
        if not de or de < data_inicio or de > data_fim:
            continue
        if (order.get('status') or '').lower() in STATUS_OK_IGNORAR:
            continue
        for item in order.get('items') or []:
            if not isinstance(item, dict):
                continue
            nome = (item.get('product_name') or item.get('name') or '').strip()
            try:
                qtd = int(round(float(item.get('quantity', 0) or 0)))
            except (TypeError, ValueError):
                qtd = 0
            if not nome or qtd <= 0:
                continue
            sku = (item.get('sku') or item.get('product_sku') or '').strip() or None
            mp = None
            if sku:
                mp = VndaProdutoMap.query.filter_by(vnda_sku=sku).first()
            if not mp:
                mp = (VndaProdutoMap.query
                      .filter(VndaProdutoMap.vnda_nome.ilike(nome))
                      .first())
            if not mp or mp.ignorar:
                continue
            if mp.estado != 'mapeado':
                continue

            # Se for Produto cesta, desempacota nos componentes (mesma logica
            # do vnda_sync._baixar_item). Multiplica qtd vendida pela qtd
            # de cada componente dentro do produto.
            componentes = _componentes_de_cesta(mp.produto) if mp.produto_id else []
            if componentes:
                n_processados += 1
                for tipo_c, item_id_c, _, qtd_no_item in componentes:
                    if not item_id_c:
                        continue
                    chave_c = (tipo_c if tipo_c != 'mp' else 'mp', item_id_c)
                    # Soma proporcional. Pra inteiros, arredonda.
                    qtd_total = int(round(qtd * float(qtd_no_item or 1.0)))
                    if qtd_total > 0:
                        vendas[chave_c] += qtd_total
                continue

            # Nao eh cesta — conta direto
            chave = None
            if mp.receita_id:
                chave = ('receita', mp.receita_id)
            elif mp.produto_id:
                chave = ('produto', mp.produto_id)
            elif mp.materia_prima_id:
                chave = ('mp', mp.materia_prima_id)
            if not chave:
                continue
            vendas[chave] += qtd
            n_processados += 1
    ULTIMA_CONSULTA_VNDA[stats_key]['n_processados'] = n_processados
    return dict(vendas), None


def parsear_lista(texto):
    """Reusa o parser do estoque_loja_lote (mesmo formato 'Nome: qtd')."""
    return svc_lote.parsear_lista(texto)


def resolver_lista(parseados, loja_id):
    """Reusa o resolver (fuzzy match + apelidos globais)."""
    return svc_lote.resolver_lista(parseados, loja_id)


def aplicar_vendas_manuais(itens_resolvidos, loja_id, data_venda, user):
    """Cria VendaManualLoja pra cada item resolvido. NAO mexe em estoque.

    Itens nao resolvidos sao ignorados (devolve em `ignorados`). Pra
    nao perder histórico, o admin pode vincular o apelido depois e
    relancar.

    Retorna {aplicados: [{nome, tipo, quantidade}], ignorados: [...]}.
    """
    if not loja_id or not data_venda:
        return {'aplicados': [], 'ignorados': [{'linha': '*', 'motivo': 'sem_loja_ou_data'}]}

    aplicados = []
    ignorados = []

    for item in itens_resolvidos:
        if item.get('erro'):
            ignorados.append({'linha': item.get('linha', '?'), 'motivo': item['erro']})
            continue
        resolvido = item.get('resolvido')
        if not resolvido or not resolvido.get('id'):
            ignorados.append({'linha': item.get('linha', '?'),
                                'nome': item.get('nome'),
                                'motivo': 'nao_resolvido'})
            continue
        try:
            qtd = int(item['quantidade'])
        except (KeyError, TypeError, ValueError):
            ignorados.append({'linha': item.get('linha', '?'), 'motivo': 'qtd_invalida'})
            continue
        if qtd <= 0:
            continue

        # Pendente NAO entra como venda (orfao no estoque, nao vinculado)
        if resolvido['tipo'] == 'pendente':
            ignorados.append({'linha': item.get('linha', '?'),
                                'nome': item.get('nome'),
                                'motivo': 'item_pendente_de_vinculacao'})
            continue

        rid = resolvido['id'] if resolvido['tipo'] == 'receita' else None
        pid = resolvido['id'] if resolvido['tipo'] == 'produto' else None
        mid = resolvido['id'] if resolvido['tipo'] == 'mp' else None

        # Dedupe — evita upload duplicado virar venda em dobro
        existente = VendaManualLoja.query.filter_by(
            loja_id=loja_id, data_venda=data_venda,
            receita_id=rid, produto_id=pid, materia_prima_id=mid,
        ).first()
        if existente:
            ignorados.append({'linha': item.get('linha', '?'),
                                'nome': resolvido['nome'],
                                'motivo': 'duplicata_ja_lancada'})
            continue

        vm = VendaManualLoja(
            loja_id=loja_id, data_venda=data_venda,
            receita_id=rid, produto_id=pid, materia_prima_id=mid,
            quantidade=qtd,
            criado_por_id=getattr(user, 'id', None),
        )
        db.session.add(vm)
        aplicados.append({
            'nome': resolvido['nome'],
            'tipo': resolvido['tipo'],
            'quantidade': qtd,
        })

    if aplicados:
        db.session.commit()
    return {'aplicados': aplicados, 'ignorados': ignorados}


def _chave_item(receita_id=None, produto_id=None, mp_id=None):
    if receita_id:
        return ('receita', receita_id)
    if produto_id:
        return ('produto', produto_id)
    if mp_id:
        return ('mp', mp_id)
    return None


def sugerir_pedido(loja_id, data_inicio=None, data_fim=None,
                    dias_cobertura=7):
    """Calcula sugestao de pedido pra uma loja num intervalo de datas.

    Soma vendas reais (Seru/VNDA via MovEstoqueLoja) + vendas manuais
    (VendaManualLoja) entre data_inicio e data_fim. Calcula media diaria
    e sugere qtd pra `dias_cobertura` dias.

    data_inicio/data_fim sao `date`. Se nao fornecidos, usa ultimos 14 dias.
    Retorna lista [{tipo, id, nome, media_diaria, estoque_atual,
                    qtd_sugerida, vendas_periodo, por_fonte}].
    `por_fonte` = {'vnda': qtd, 'seru': qtd, 'manual': qtd}
    """
    if not loja_id:
        return {'itens': [], 'aviso_vnda': None}
    if data_fim is None:
        data_fim = hoje()
    if data_inicio is None:
        data_inicio = data_fim - timedelta(days=14)
    if data_inicio > data_fim:
        data_inicio, data_fim = data_fim, data_inicio
    dias_periodo = max(1, (data_fim - data_inicio).days + 1)

    # 1. Vendas reais via MovEstoqueLoja
    vendas_por_item = defaultdict(int)  # (tipo, id) → qtd_total
    fontes_por_item = defaultdict(set)
    # 1. Seru via MovEstoqueLoja (sync funciona historico via cron). VNDA
    # NAO vem daqui — buscamos direto da API embaixo (cron VNDA so processa
    # o dia corrente, entao MovEstoqueLoja nao tem retroativo).
    from datetime import datetime
    dt_inicio = datetime.combine(data_inicio, time.min)
    dt_fim = datetime.combine(data_fim, time.max)
    por_fonte_item = defaultdict(lambda: defaultdict(int))  # {(tipo,id): {fonte: qtd}}
    # Inclui `venda_seru_estorno` pra subtrair pedidos cancelados — senao
    # a media diaria fica inflada por vendas grandes que foram estornadas.
    SERU_TIPOS = ('venda_seru', 'venda_seru_sem_estoque', 'venda_seru_estorno')
    movs = (db.session.query(MovEstoqueLoja, EstoqueLoja)
            .join(EstoqueLoja, MovEstoqueLoja.estoque_loja_id == EstoqueLoja.id)
            .filter(EstoqueLoja.loja_id == loja_id,
                    MovEstoqueLoja.tipo.in_(SERU_TIPOS),
                    MovEstoqueLoja.data >= dt_inicio,
                    MovEstoqueLoja.data <= dt_fim)
            .all())
    for mov, el in movs:
        chave = _chave_item(el.receita_id, el.produto_id, el.materia_prima_id)
        if not chave:
            continue
        qtd = int(mov.quantidade or 0)
        # Estornos subtraem do total — a venda original ja contou positivo,
        # o estorno cancela essa contribuicao.
        if mov.tipo == 'venda_seru_estorno':
            qtd = -qtd
        vendas_por_item[chave] += qtd
        fontes_por_item[chave].add('seru')
        por_fonte_item[chave]['seru'] += qtd

    # 1b. VNDA aposentado (canal e-commerce desligado em 06/2026 — operacao
    # 100% no sistema proprio). NAO batemos mais na API VNDA nesta sugestao:
    # o painel de producao passou a prever pela demanda real (PedidoLoja) e
    # esta sugestao por loja usa Seru + manual. `_agregar_vendas_vnda_api`
    # segue existindo pra leitura de historico em /pdv/itens-vendidos.
    aviso_vnda = None

    # 2. Vendas manuais (por data_venda — Date, nao datetime)
    manuais = VendaManualLoja.query.filter(
        VendaManualLoja.loja_id == loja_id,
        VendaManualLoja.data_venda >= data_inicio,
        VendaManualLoja.data_venda <= data_fim,
    ).all()
    for vm in manuais:
        chave = _chave_item(vm.receita_id, vm.produto_id, vm.materia_prima_id)
        if not chave:
            continue
        qtd = int(vm.quantidade or 0)
        vendas_por_item[chave] += qtd
        fontes_por_item[chave].add('manual')
        por_fonte_item[chave]['manual'] += qtd

    if not vendas_por_item:
        return {'itens': [], 'aviso_vnda': aviso_vnda}

    # 3. Estoque atual da loja por chave
    estoque_por_item = {}
    for el in EstoqueLoja.query.filter_by(loja_id=loja_id).all():
        chave = _chave_item(el.receita_id, el.produto_id, el.materia_prima_id)
        if chave:
            estoque_por_item[chave] = el.quantidade or 0

    # 4. Resolve nomes
    nome_por_chave = {}
    receitas_ids = [k[1] for k in vendas_por_item if k[0] == 'receita']
    produtos_ids = [k[1] for k in vendas_por_item if k[0] == 'produto']
    mps_ids = [k[1] for k in vendas_por_item if k[0] == 'mp']
    if receitas_ids:
        for r in Receita.query.filter(Receita.id.in_(receitas_ids)).all():
            nome_por_chave[('receita', r.id)] = r.nome
    if produtos_ids:
        for p in Produto.query.filter(Produto.id.in_(produtos_ids)).all():
            nome_por_chave[('produto', p.id)] = p.nome
    if mps_ids:
        for m in MateriaPrima.query.filter(MateriaPrima.id.in_(mps_ids)).all():
            nome_por_chave[('mp', m.id)] = m.nome

    # 5. Monta sugestao
    out = []
    for chave, total_vendas in vendas_por_item.items():
        tipo, item_id = chave
        media = total_vendas / dias_periodo
        estoque_atual = estoque_por_item.get(chave, 0)
        ideal = math.ceil(media * dias_cobertura)
        qtd_sugerida = max(0, ideal - estoque_atual)
        out.append({
            'tipo': tipo,
            'id': item_id,
            'nome': nome_por_chave.get(chave, '?'),
            'media_diaria': round(media, 2),
            'vendas_periodo': total_vendas,
            'estoque_atual': estoque_atual,
            'ideal_cobertura': ideal,
            'qtd_sugerida': qtd_sugerida,
            'fontes': sorted(fontes_por_item.get(chave, [])),
            'por_fonte': dict(por_fonte_item.get(chave, {})),
        })
    out.sort(key=lambda x: -x['media_diaria'])
    return {'itens': out, 'aviso_vnda': aviso_vnda}


# ── Upload / template Excel ──

def exportar_sugestao_xlsx(loja, sugestao_itens, data_inicio, data_fim,
                            dias_cobertura):
    """Gera xlsx da tela de sugestao de pedido. 1 aba com tabela completa
    + cabecalho com periodo. Retorna bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sugestão'

    # Cabecalho
    ws['A1'] = f'Sugestão de Pedido — {loja.nome}'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:I1')
    ws['A2'] = f'Período: {data_inicio} a {data_fim} ({(data_fim - data_inicio).days + 1} dias) · Cobertura: {dias_cobertura} dia(s)'
    ws['A2'].font = Font(italic=True, color='666666')
    ws.merge_cells('A2:I2')

    # Header tabela
    headers = ['Item', 'PDV manual', 'VNDA', 'Seru', 'Total no período',
                'Média/dia', 'Estoque atual', 'Ideal cobertura', 'Pedir']
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2c3e50')
        c.alignment = Alignment(horizontal='center')

    # Linhas
    for i, s in enumerate(sugestao_itens, start=5):
        ws.cell(row=i, column=1, value=s['nome'])
        ws.cell(row=i, column=2, value=s.get('por_fonte', {}).get('manual', 0) or 0)
        ws.cell(row=i, column=3, value=s.get('por_fonte', {}).get('vnda', 0) or 0)
        ws.cell(row=i, column=4, value=s.get('por_fonte', {}).get('seru', 0) or 0)
        ws.cell(row=i, column=5, value=s.get('vendas_periodo', 0)).font = Font(bold=True)
        ws.cell(row=i, column=6, value=s.get('media_diaria', 0))
        ws.cell(row=i, column=7, value=s.get('estoque_atual', 0))
        ws.cell(row=i, column=8, value=s.get('ideal_cobertura', 0))
        c_pedir = ws.cell(row=i, column=9, value=s.get('qtd_sugerida', 0))
        c_pedir.font = Font(bold=True)
        if (s.get('qtd_sugerida') or 0) > 0:
            c_pedir.fill = PatternFill('solid', fgColor='e8f5e9')

    # Larguras
    larguras = {'A': 38, 'B': 12, 'C': 10, 'D': 10, 'E': 16, 'F': 12,
                'G': 14, 'H': 14, 'I': 10}
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w

    # Totais no rodape (linha apos os itens)
    rod = len(sugestao_itens) + 6
    ws.cell(row=rod, column=1, value='TOTAL').font = Font(bold=True)
    for col_idx, key in enumerate(
        [None, 'manual', 'vnda', 'seru', 'vendas_periodo', None,
         'estoque_atual', 'ideal_cobertura', 'qtd_sugerida'], start=1):
        if not key:
            continue
        if key in ('manual', 'vnda', 'seru'):
            total = sum((s.get('por_fonte', {}).get(key) or 0) for s in sugestao_itens)
        else:
            total = sum((s.get(key) or 0) for s in sugestao_itens)
        c = ws.cell(row=rod, column=col_idx, value=total)
        c.font = Font(bold=True)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def gerar_template_xlsx(loja):
    """Gera planilha modelo pro admin preencher e fazer upload.

    3 colunas: Data, Produto, Quantidade. Inclui linhas de exemplo com
    receitas do catalogo da padaria pra ele so copiar/preencher.
    Retorna bytes (xlsx).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Vendas'

    # Cabecalho
    headers = ['Data', 'Produto', 'Quantidade']
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2c3e50')
        cell.alignment = Alignment(horizontal='center')
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 12

    # Linhas de exemplo: pega ate 20 receitas + 10 produtos
    exemplos_nomes = [r.nome for r in
                      Receita.query.filter(Receita.arquivada_em.is_(None))
                      .order_by(Receita.categoria, Receita.nome).limit(30).all()]
    if not exemplos_nomes:
        exemplos_nomes = ['Croissant Tradicional', 'Pao Frances', 'Sourdough']

    data_exemplo = hoje().isoformat()
    for i, nome in enumerate(exemplos_nomes, start=2):
        ws.cell(row=i, column=1, value=data_exemplo)
        ws.cell(row=i, column=2, value=nome)
        ws.cell(row=i, column=3, value=0)
        ws.cell(row=i, column=1).number_format = 'YYYY-MM-DD'

    # Aba de instrucoes
    ws_help = wb.create_sheet('Como usar')
    instrucoes = [
        f'Vendas manuais — Loja {loja.nome}',
        '',
        '1. Coluna Data: YYYY-MM-DD (ex: 2026-04-15) ou DD/MM/YYYY.',
        '2. Coluna Produto: nome igual ao catalogo. Apelidos salvos funcionam.',
        '3. Coluna Quantidade: numero inteiro > 0. Use 0 ou apague a linha pra ignorar.',
        '4. Pode misturar varias datas e produtos na mesma planilha.',
        '5. Linhas com quantidade 0 sao ignoradas.',
        '6. Sistema NAO baixa estoque — so registra historico pra sugerir pedido.',
        '',
        'Exemplo:',
        '2026-04-01  |  Croissant Tradicional  |  25',
        '2026-04-01  |  Pao Frances  |  80',
        '2026-04-02  |  Croissant Tradicional  |  30',
        '...',
        '',
        'Apos upload, confira o preview e clique em "Aplicar".',
    ]
    for i, linha in enumerate(instrucoes, start=1):
        ws_help.cell(row=i, column=1, value=linha)
    ws_help.column_dimensions['A'].width = 80

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def parsear_xlsx(file_storage, loja_id):
    """Le xlsx (file_storage do Flask) e retorna lista de itens parseados
    pra resolver_lista_xlsx aplicar. Aceita data como string YYYY-MM-DD ou
    DD/MM/YYYY, ou como datetime/date (Excel as vezes joga assim).

    Retorna [{linha_n, data_venda, nome, quantidade, erro?}].
    Itens com erro nao bloqueiam — entram em ignorados depois.
    """
    from openpyxl import load_workbook
    try:
        wb = load_workbook(file_storage, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return [{'linha_n': 0, 'erro': f'arquivo invalido: {exc}'}]

    ws = wb.active  # primeira aba
    rows = ws.iter_rows(min_row=2, values_only=True)  # pula cabecalho
    out = []
    for n, row in enumerate(rows, start=2):
        if not row or all(c is None or c == '' for c in row):
            continue
        try:
            data_raw = row[0]
            nome_raw = row[1]
            qtd_raw = row[2]
        except IndexError:
            out.append({'linha_n': n, 'erro': 'colunas insuficientes (esperado: Data, Produto, Qtd)'})
            continue

        # Parse data
        data_venda = None
        if isinstance(data_raw, (datetime, date)):
            data_venda = data_raw.date() if isinstance(data_raw, datetime) else data_raw
        elif isinstance(data_raw, str):
            data_str = data_raw.strip()
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d/%m/%y'):
                try:
                    data_venda = datetime.strptime(data_str, fmt).date()
                    break
                except ValueError:
                    continue
        if not data_venda:
            out.append({'linha_n': n, 'data_raw': data_raw,
                         'erro': 'data invalida (use YYYY-MM-DD ou DD/MM/YYYY)'})
            continue

        nome = (str(nome_raw) if nome_raw is not None else '').strip()
        if not nome:
            out.append({'linha_n': n, 'erro': 'nome vazio'})
            continue

        try:
            qtd = int(float(qtd_raw)) if qtd_raw is not None else 0
        except (TypeError, ValueError):
            out.append({'linha_n': n, 'nome': nome,
                         'erro': f'quantidade invalida: {qtd_raw}'})
            continue
        if qtd <= 0:
            continue  # linhas com 0 sao silenciosamente puladas

        out.append({'linha_n': n, 'data_venda': data_venda,
                     'nome': nome, 'quantidade': qtd})
    return out


def aplicar_vendas_xlsx(itens_parseados, loja_id, user):
    """Resolve cada nome via fuzzy + apelidos, cria VendaManualLoja.

    itens_parseados vem de parsear_xlsx (lista com data_venda + nome + qtd).
    Resolve uma vez por nome unico (cacheado) pra economizar query.

    Retorna {aplicados: [...], ignorados: [...], total_linhas, datas_unicas}.
    """
    if not loja_id:
        return {'aplicados': [], 'ignorados': [], 'total_linhas': 0,
                'datas_unicas': []}

    # Pre-resolve cada nome unico de uma vez
    nomes_unicos = sorted({it['nome'] for it in itens_parseados if it.get('nome')})
    parseados_p_resolver = [{'linha': nome, 'nome': nome, 'quantidade': 1}
                            for nome in nomes_unicos]
    resolvidos_lista = svc_lote.resolver_lista(parseados_p_resolver, loja_id)
    resolvidos_por_nome = {r['nome']: r.get('resolvido') for r in resolvidos_lista}

    aplicados = []
    ignorados = []
    datas_set = set()

    for it in itens_parseados:
        if it.get('erro'):
            ignorados.append({'linha_n': it['linha_n'], 'motivo': it['erro']})
            continue
        nome = it['nome']
        resolvido = resolvidos_por_nome.get(nome)
        if not resolvido or not resolvido.get('id'):
            ignorados.append({'linha_n': it['linha_n'], 'nome': nome,
                                'motivo': 'nao_resolvido'})
            continue
        if resolvido['tipo'] == 'pendente':
            ignorados.append({'linha_n': it['linha_n'], 'nome': nome,
                                'motivo': 'item_pendente_de_vinculacao'})
            continue

        rid = resolvido['id'] if resolvido['tipo'] == 'receita' else None
        pid = resolvido['id'] if resolvido['tipo'] == 'produto' else None
        mid = resolvido['id'] if resolvido['tipo'] == 'mp' else None

        # Dedupe: se ja existe lancamento exato pra (loja, data, item),
        # nao cria de novo. Cobre upload duplicado do mesmo xlsx.
        existente = VendaManualLoja.query.filter_by(
            loja_id=loja_id, data_venda=it['data_venda'],
            receita_id=rid, produto_id=pid, materia_prima_id=mid,
        ).first()
        if existente:
            ignorados.append({
                'linha_n': it['linha_n'], 'nome': resolvido['nome'],
                'motivo': 'duplicata_ja_lancada',
                'detalhe': f'qtd existente: {existente.quantidade}',
            })
            continue

        vm = VendaManualLoja(
            loja_id=loja_id, data_venda=it['data_venda'],
            receita_id=rid, produto_id=pid, materia_prima_id=mid,
            quantidade=it['quantidade'],
            criado_por_id=getattr(user, 'id', None),
        )
        db.session.add(vm)
        aplicados.append({
            'linha_n': it['linha_n'],
            'data': it['data_venda'].isoformat(),
            'nome': resolvido['nome'],
            'quantidade': it['quantidade'],
        })
        datas_set.add(it['data_venda'])

    if aplicados:
        db.session.commit()

    return {
        'aplicados': aplicados,
        'ignorados': ignorados,
        'total_linhas': len(itens_parseados),
        'datas_unicas': sorted(datas_set),
    }

"""Exportacao XLSX da tabela de precos (02/07/2026, pedido do dono).

Uma aba com TODOS os itens vendaveis — receitas, produtos simples e cestas —
no layout pedido: PRODUTO | CUSTO | PRECO LOJA | PRECO SITE | PRECO INTERNO |
ATACADO (+ TIPO e CATEGORIA como colunas de apoio no fim, pra filtrar).

Fontes (as MESMAS da tela /receitas/precos):
- custo: custos.calcular_custos_receitas / calcular_custos_produtos;
- atacado: Receita.preco_venda (rotulado "Atacado" na UI) e
  Produto.preco_atacado — nomes de coluna diferentes por historico.
Receitas arquivadas e produtos inativos ficam de fora (mesma regra da tela).
"""
import io

from app.models import Produto, Receita
from app.services.custos import calcular_custos_produtos, calcular_custos_receitas


def _linhas_precos():
    """Coleta as linhas na ordem: receitas, produtos simples, cestas —
    cada uma (nome, tipo, categoria, custo, loja, site, interno, atacado)."""
    res_custos = calcular_custos_receitas()
    custos_receita = res_custos['custos']          # {nome: custo unitario}
    custos_produto = calcular_custos_produtos(
        custos_receita, res_custos['mp_info'])     # {nome: custo}

    linhas = []
    for r in (Receita.query.filter(Receita.arquivada_em.is_(None))
              .order_by(Receita.categoria, Receita.nome).all()):
        linhas.append((r.nome, 'Receita', r.categoria or '',
                       custos_receita.get(r.nome),
                       r.preco_loja, r.preco_site, r.preco_interno,
                       r.preco_venda))
    produtos = (Produto.query.filter_by(ativo=True)
                .order_by(Produto.categoria, Produto.nome).all())
    simples = [p for p in produtos if not p.itens]
    cestas = [p for p in produtos if p.itens]
    for grupo, tipo in ((simples, 'Produto'), (cestas, 'Cesta')):
        for p in grupo:
            linhas.append((p.nome, tipo, p.categoria or '',
                           custos_produto.get(p.nome),
                           p.preco_loja, p.preco_site, p.preco_interno,
                           p.preco_atacado))
    return linhas


def gerar_xlsx_precos():
    """Gera o .xlsx (bytes) da tabela de precos."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    from app.utils import hoje

    wb = Workbook()
    ws = wb.active
    ws.title = 'Precos'

    ws['A1'] = 'Tabela de preços — O Pão Padaria Artesanal'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = 'Gerada em %s. Custo = referência do custeio atual '\
               '(ingredientes + embalagem); vazio = sem ficha de custo.' \
               % hoje().strftime('%d/%m/%Y')
    ws['A2'].font = Font(italic=True, size=9, color='666666')

    cols = ['PRODUTO', 'CUSTO', 'PRECO LOJA', 'PRECO SITE', 'PRECO INTERNO',
            'ATACADO', 'TIPO', 'CATEGORIA']
    larguras = [42, 12, 12, 12, 14, 12, 10, 18]
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='37474F', end_color='37474F',
                              fill_type='solid')
    hrow = 4
    for c, nome in enumerate(cols, start=1):
        cell = ws.cell(row=hrow, column=c, value=nome)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = larguras[c - 1]

    r = hrow + 1
    for nome, tipo, categoria, custo, loja, site, interno, atacado in \
            _linhas_precos():
        ws.cell(row=r, column=1, value=nome)
        for c, valor in ((2, custo), (3, loja), (4, site), (5, interno),
                         (6, atacado)):
            if valor is not None:
                cell = ws.cell(row=r, column=c, value=round(float(valor), 2))
                cell.number_format = 'R$ #,##0.00'
        ws.cell(row=r, column=7, value=tipo)
        ws.cell(row=r, column=8, value=categoria)
        r += 1

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = 'A4:H%d' % max(r - 1, hrow)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

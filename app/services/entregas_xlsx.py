"""XLSX da aba Produtos do /entregas (08/08/2026, pedido do dono na véspera
do Dia dos Pais: "Preciso poder gerar essa lista em xlsx").

UMA aba só, espelho fiel da tela: "Vendidos no dia" (como vendido — cesta é
1 linha) e, ABAIXO, "A produzir" (cestas explodidas em componentes, com
unidade e de qual cesta cada um vem). Ajustes do dono 08/08/2026 após o 1º
uso real: os dois blocos na MESMA aba (a segunda aba passava despercebida no
celular) e **SEM valores em R$** — a planilha circula com a equipe de
montagem ("ninguém precisa saber dos valores, só eu"); dinheiro fica nas
telas do dono.

Recebe o dict de `entregas.routes._produtos_do_dia` — o MESMO motor que
alimenta a aba; o servidor re-agrega e nunca confia no estado do navegador.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FONT = Font(bold=True, color='FFFFFF')
_HEADER_FILL = PatternFill(start_color='37474F', end_color='37474F',
                           fill_type='solid')
_TITULO_FONT = Font(bold=True, size=13)
_BOLD = Font(bold=True)


def _titulo(ws, texto):
    ws.append([texto])
    ws.cell(row=ws.max_row, column=1).font = _TITULO_FONT


def _header(ws, cols):
    ws.append(cols)
    linha = ws.max_row
    for i in range(1, len(cols) + 1):
        cell = ws.cell(row=linha, column=i)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL


def _negrito(ws):
    for cell in ws[ws.max_row]:
        cell.font = _BOLD


def gerar_xlsx_produtos_dia(dados):
    """Gera o .xlsx (bytes). `dados` = dict do `_produtos_do_dia`."""
    janelas = dados.get('janelas') or []
    sub = 'Data: %s · %s · %d pedido(s)' % (
        dados.get('data') or '',
        ('Janelas: ' + ', '.join(janelas)) if janelas else 'Todas as janelas',
        dados.get('total_pedidos') or 0)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Produtos do dia'
    for i, larg in enumerate((42, 12, 10, 46), start=1):
        ws.column_dimensions[get_column_letter(i)].width = larg

    # ── Bloco 1: vendidos, como o cliente comprou (sem R$) ────────────────
    _titulo(ws, 'Vendidos no dia')
    ws.append([sub])
    ws.append([])
    _header(ws, ['Produto', 'Qtd'])
    for v in dados.get('vendidos') or []:
        ws.append([v.get('nome'), v.get('quantidade') or 0])
    ws.append([])
    ws.append(['TOTAL', dados.get('total_itens_vendidos') or 0])
    _negrito(ws)

    # ── Bloco 2: a produzir, cestas explodidas — ABAIXO, mesma aba ────────
    ws.append([])
    ws.append([])
    _titulo(ws, 'A produzir (cestas explodidas)')
    ws.append([])
    _header(ws, ['Produto', 'Qtd', 'Unidade', 'Componente de'])
    for p in dados.get('producao') or []:
        ws.append([p.get('nome'), p.get('quantidade') or 0,
                   p.get('unidade') or 'un',
                   ', '.join(p.get('componente_de') or [])])
    ws.append([])
    # Totais POR UNIDADE — 300 g + 12 un não são somáveis (mesma regra da tela).
    tot = dados.get('totais_producao_por_unidade') or {}
    _ordem = {'un': 0, 'g': 1, 'kg': 2, 'ml': 3, 'l': 4}
    partes = ['%s %s' % (tot[u], u)
              for u in sorted(tot, key=lambda u: _ordem.get(u, 99))]
    ws.append(['TOTAL', ' · '.join(partes) if partes else '0'])
    _negrito(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

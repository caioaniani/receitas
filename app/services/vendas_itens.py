"""Agregacao de itens vendidos da Seru.

Para cada produto vendido num intervalo, calcula:
- quantidade total vendida
- faturamento total
- numero de pedidos distintos
- match no catalogo local (Receita ou Produto), por fuzzy
- estado do mapeamento Seru (VendaMapa canal=seru): mapeado/ignorado/pendente/sem_map

Filtros: intervalo de datas BRT e (opcional) nome da loja Seru
(o campo 'company.name' do pedido — que e o que a Seru chama de loja).
"""
import unicodedata

from app.models import Produto, Receita, VendaMapa
from app.services import seru


def _ascii(s):
    if not s:
        return ''
    nf = unicodedata.normalize('NFD', s)
    return ''.join(c for c in nf if unicodedata.category(c) != 'Mn').lower().strip()


def _carregar_catalogo():
    # ativas(): a coluna "match no catalogo" orienta o Vincular dos
    # mapeamentos Seru — sugerir receita arquivada induzia a criar VendaMapa
    # pra ela (varredura 19/07/2026). Produto ja filtrava.
    receitas = [(r.id, r.nome, _ascii(r.nome)) for r in Receita.ativas().all()]
    produtos = [(p.id, p.nome, _ascii(p.nome))
                for p in Produto.query.filter_by(ativo=True).all()]
    return receitas, produtos


def _match_local(nome, receitas, produtos):
    """Retorna {'tipo': 'receita'|'produto', 'id', 'nome', 'kind': 'exato'|'fuzzy'}
    ou None se nao houver match razoavel."""
    alvo = _ascii(nome)
    if not alvo:
        return None
    # 1. exato
    for rid, rnome, rasc in receitas:
        if rasc == alvo:
            return {'tipo': 'receita', 'id': rid, 'nome': rnome, 'kind': 'exato'}
    for pid, pnome, pasc in produtos:
        if pasc == alvo:
            return {'tipo': 'produto', 'id': pid, 'nome': pnome, 'kind': 'exato'}
    # 2. substring
    for rid, rnome, rasc in receitas:
        if alvo in rasc or rasc in alvo:
            return {'tipo': 'receita', 'id': rid, 'nome': rnome, 'kind': 'fuzzy'}
    for pid, pnome, pasc in produtos:
        if alvo in pasc or pasc in alvo:
            return {'tipo': 'produto', 'id': pid, 'nome': pnome, 'kind': 'fuzzy'}
    return None


def _nome_loja(pedido):
    """Extrai o nome da loja Seru do pedido (campo 'company.name' tipicamente)."""
    c = pedido.get('company')
    if isinstance(c, dict):
        return (c.get('name') or c.get('label') or '').strip()
    if isinstance(c, str):
        return c.strip()
    return ''


def agregar_itens(data_inicial, data_final, loja_seru=None,
                  expandir_dias_frente=0):
    """Pega pedidos da Seru no intervalo, agrega por nome de produto.

    Retorna:
        {
          'inicio': iso, 'fim': iso, 'loja': str|None,
          'total_pedidos': N, 'total_itens_vendidos': N,
          'faturamento_total': float,
          'produtos': [{nome, qtd, faturamento, n_pedidos, pct_faturamento, match}],
          'sem_match_count': N,
          'lojas_no_intervalo': [lojaA, lojaB, ...],  # pra preencher dropdown
        }
    """
    receitas, produtos = _carregar_catalogo()

    pedidos = seru.listar_pedidos_completo(
        data_inicial, data_final, expandir_dias_frente=expandir_dias_frente)

    # Filtra por createdAt no intervalo BRT + (opcional) loja
    lojas_vistas = set()
    pedidos_filtrados = []
    for p in pedidos:
        if not isinstance(p, dict):
            continue
        if p.get('canceledAt'):
            continue
        d = seru.data_local(p.get('createdAt'))
        if not d or not (data_inicial <= d <= data_final):
            continue
        ln = _nome_loja(p)
        if ln:
            lojas_vistas.add(ln)
        if loja_seru and ln != loja_seru:
            continue
        pedidos_filtrados.append(p)

    # Agrega por nome do produto Seru
    agg = {}  # nome -> {qtd, faturamento, n_pedidos (set de ids), sku}
    for p in pedidos_filtrados:
        pid = p.get('id') or p.get('orderNumber') or p.get('code')
        for it in seru.extrair_itens(p):
            if it['cancelado']:
                continue
            nome = it['nome']
            if nome not in agg:
                agg[nome] = {'qtd': 0.0, 'faturamento': 0.0,
                             'pedidos': set(), 'sku': it['sku']}
            agg[nome]['qtd'] += it['qtd']
            agg[nome]['faturamento'] += it['total']
            if pid is not None:
                agg[nome]['pedidos'].add(pid)

    faturamento_total = sum(v['faturamento'] for v in agg.values())
    total_itens = sum(v['qtd'] for v in agg.values())

    # Index dos VendaMapa(canal=seru) pra mostrar estado nas linhas.
    maps = {m.nome_externo: m for m in VendaMapa.query.filter(
        VendaMapa.canal == 'seru',
        VendaMapa.nome_externo.in_(list(agg.keys()))).all()}

    produtos_lista = []
    sem_match = 0
    pendentes = 0
    for nome, v in agg.items():
        match = _match_local(nome, receitas, produtos)
        if not match:
            sem_match += 1
        # Estado do mapeamento Seru (autoritativo pra auto-baixa).
        m = maps.get(nome)
        if m:
            estado_map = m.estado  # mapeado/ignorado/pendente
            mapeado_para = {
                'tipo': 'receita' if m.receita_id else ('produto' if m.produto_id else None),
                'id': m.receita_id or m.produto_id,
                'nome': m.alvo_nome,
            } if m.estado == 'mapeado' else None
            map_id = m.id
            fator = float(m.fator_quantidade or 1.0)
        else:
            estado_map = 'sem_map'  # ainda nao foi visto numa sync
            mapeado_para = None
            map_id = None
            fator = 1.0
        if estado_map in ('pendente', 'sem_map'):
            pendentes += 1
        produtos_lista.append({
            'nome': nome,
            'sku': v['sku'],
            'qtd': v['qtd'],
            'faturamento': round(v['faturamento'], 2),
            'n_pedidos': len(v['pedidos']),
            'pct_faturamento': round(100 * v['faturamento'] / faturamento_total, 1)
                if faturamento_total else 0.0,
            'match': match,  # palpite por fuzzy local (sugestao)
            'estado_map': estado_map,
            'mapeado_para': mapeado_para,
            'map_id': map_id,
            'fator': fator,
        })
    produtos_lista.sort(key=lambda x: x['faturamento'], reverse=True)

    return {
        'inicio': data_inicial.isoformat(),
        'fim': data_final.isoformat(),
        'loja': loja_seru,
        'total_pedidos': len(pedidos_filtrados),
        'total_itens_vendidos': round(total_itens, 2),
        'faturamento_total': round(faturamento_total, 2),
        'produtos': produtos_lista,
        'sem_match_count': sem_match,  # mantido por compat
        'pendentes_count': pendentes,
        'lojas_no_intervalo': sorted(lojas_vistas),
    }


def _linhas_produtos(pedidos, receitas, produtos, maps):
    """Agrega uma lista de pedidos por NOME de produto Seru.

    Retorna (linhas, faturamento_total, total_itens) — `linhas` tem a MESMA forma
    de `agregar_itens` (nome/sku/qtd/faturamento/n_pedidos/pct/match/estado_map/
    mapeado_para/map_id/fator). `maps` = {nome_externo: VendaMapa} pra estado do
    mapeamento (consultado 1x pelo caller pra todos os nomes)."""
    agg = {}
    for p in pedidos:
        pid = p.get('id') or p.get('orderNumber') or p.get('code')
        for it in seru.extrair_itens(p):
            if it['cancelado']:
                continue
            nome = it['nome']
            e = agg.setdefault(nome, {'qtd': 0.0, 'faturamento': 0.0,
                                      '_peds': set(), 'sku': it['sku']})
            e['qtd'] += it['qtd']
            e['faturamento'] += it['total']
            if pid is not None:
                e['_peds'].add(pid)
    for e in agg.values():
        e['n_pedidos'] = len(e.pop('_peds'))
    return montar_linhas(agg, receitas, produtos, maps)


def montar_linhas(agg, receitas, produtos, maps):
    """Constroi as linhas do relatorio a partir de um agregado ja pronto.

    `agg` = {nome: {'qtd', 'faturamento', 'n_pedidos', 'sku'}}. Fonte unica da
    forma da linha (nome/sku/qtd/faturamento/n_pedidos/pct/match/estado_map/
    mapeado_para/map_id/fator) — usada tanto pela agregacao AO VIVO (`_linhas_
    produtos`, a partir dos pedidos da API) quanto pela leitura do BANCO
    (`vendas_diarias`), pra as duas nunca divergirem. `maps` = {nome_externo:
    VendaMapa}. Retorna (linhas ordenadas por faturamento, fat_total, itens)."""
    faturamento_total = sum(v['faturamento'] for v in agg.values())
    total_itens = sum(v['qtd'] for v in agg.values())
    linhas = []
    for nome, v in agg.items():
        match = _match_local(nome, receitas, produtos)
        m = maps.get(nome)
        if m:
            estado_map = m.estado
            mapeado_para = {
                'tipo': 'receita' if m.receita_id else ('produto' if m.produto_id else None),
                'id': m.receita_id or m.produto_id,
                'nome': m.alvo_nome,
            } if m.estado == 'mapeado' else None
            map_id = m.id
            fator = float(m.fator_quantidade or 1.0)
        else:
            estado_map = 'sem_map'
            mapeado_para = None
            map_id = None
            fator = 1.0
        linhas.append({
            'nome': nome,
            'sku': v.get('sku'),
            'qtd': round(v['qtd'], 2),
            'faturamento': round(v['faturamento'], 2),
            'n_pedidos': v.get('n_pedidos', 0),
            'pct_faturamento': round(100 * v['faturamento'] / faturamento_total, 1)
                if faturamento_total else 0.0,
            'match': match,
            'estado_map': estado_map,
            'mapeado_para': mapeado_para,
            'map_id': map_id,
            'fator': fator,
        })
    linhas.sort(key=lambda x: x['faturamento'], reverse=True)
    return linhas, round(faturamento_total, 2), round(total_itens, 2)


def agregar_itens_por_loja(data_inicial, data_final, expandir_dias_frente=0):
    """Como `agregar_itens`, mas SEPARADO POR LOJA (company.name do Seru).

    Bate na API UMA vez e reparte os pedidos por loja. Retorna:
        {
          'inicio', 'fim', 'total_pedidos', 'total_itens_vendidos',
          'faturamento_total',
          'lojas': [{'loja', 'total_pedidos', 'total_itens', 'faturamento',
                     'produtos': [...]}],           # uma entrada por loja
          'consolidado': [...],                      # todas as lojas juntas
          'lojas_no_intervalo': [...],
        }
    """
    from collections import defaultdict

    receitas, produtos = _carregar_catalogo()
    pedidos = seru.listar_pedidos_completo(
        data_inicial, data_final, expandir_dias_frente=expandir_dias_frente)

    por_loja = defaultdict(list)
    todos = []
    lojas_vistas = set()
    for p in pedidos:
        if not isinstance(p, dict) or p.get('canceledAt'):
            continue
        d = seru.data_local(p.get('createdAt'))
        if not d or not (data_inicial <= d <= data_final):
            continue
        ln = _nome_loja(p) or '(sem loja)'
        lojas_vistas.add(ln)
        por_loja[ln].append(p)
        todos.append(p)

    # Index de VendaMapa pra TODOS os nomes de uma vez (1 query).
    nomes = set()
    for p in todos:
        for it in seru.extrair_itens(p):
            if not it['cancelado']:
                nomes.add(it['nome'])
    maps = {}
    if nomes:
        maps = {m.nome_externo: m for m in VendaMapa.query.filter(
            VendaMapa.canal == 'seru',
            VendaMapa.nome_externo.in_(list(nomes))).all()}

    lojas_out = []
    for ln in sorted(por_loja):
        peds = por_loja[ln]
        linhas, fat, itens = _linhas_produtos(peds, receitas, produtos, maps)
        lojas_out.append({
            'loja': ln,
            'total_pedidos': len(peds),
            'total_itens': itens,
            'faturamento': fat,
            'produtos': linhas,
        })

    cons_linhas, cons_fat, cons_itens = _linhas_produtos(
        todos, receitas, produtos, maps)
    pendentes = sum(1 for p in cons_linhas
                    if p['estado_map'] in ('pendente', 'sem_map'))

    return {
        'inicio': data_inicial.isoformat(),
        'fim': data_final.isoformat(),
        'total_pedidos': len(todos),
        'total_itens_vendidos': cons_itens,
        'faturamento_total': cons_fat,
        'pendentes_count': pendentes,
        'lojas': lojas_out,
        'consolidado': cons_linhas,
        'lojas_no_intervalo': sorted(lojas_vistas),
    }


def _estado_map_label(estado, mapeado_para):
    if estado == 'mapeado' and mapeado_para:
        return 'Mapeado → %s' % (mapeado_para.get('nome') or '?')
    return {'ignorado': 'Ignorado', 'pendente': 'Pendente',
            'sem_map': 'Nao visto'}.get(estado, estado or '')


def gerar_xlsx_itens_por_loja(dados):
    """Gera um .xlsx (bytes) com UMA ABA POR LOJA + aba 'Consolidado'.

    Colunas: Produto | SKU | Unidades | Faturamento | Nº Pedidos | % Fat |
    Vinculo no sistema. Recebe o dict de `agregar_itens_por_loja`."""
    import io
    import re

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='37474F', end_color='37474F',
                              fill_type='solid')
    titulo_font = Font(bold=True, size=13)
    cols = ['Produto', 'SKU', 'Unidades', 'Faturamento', 'Nº Pedidos',
            '% Fat', 'Vinculo no sistema']
    larguras = [40, 16, 12, 16, 12, 10, 34]

    usados = set()

    def _titulo_aba(nome):
        # Excel: <=31 chars, sem : \\ / ? * [ ]; garante unicidade.
        base = re.sub(r'[:\\/?*\[\]]', '-', nome)[:31] or 'Loja'
        t = base
        i = 2
        while t.lower() in usados:
            suf = ' (%d)' % i
            t = base[:31 - len(suf)] + suf
            i += 1
        usados.add(t.lower())
        return t

    def _escrever_aba(ws, linhas, titulo, totais=None):
        ws['A1'] = titulo
        ws['A1'].font = titulo_font
        periodo = 'Periodo: %s a %s' % (dados['inicio'], dados['fim'])
        if totais:
            periodo += '   |   Pedidos: %s   |   Unidades: %s   |   Faturamento: R$ %.2f' % (
                totais.get('pedidos', 0), totais.get('itens', 0),
                totais.get('faturamento', 0.0))
        ws['A2'] = periodo
        ws['A2'].font = Font(italic=True, size=9, color='666666')
        hrow = 4
        for c, nome in enumerate(cols, start=1):
            cell = ws.cell(row=hrow, column=c, value=nome)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        r = hrow + 1
        for p in linhas:
            ws.cell(row=r, column=1, value=p['nome'])
            ws.cell(row=r, column=2, value=p.get('sku') or '')
            ws.cell(row=r, column=3, value=p['qtd'])
            fcell = ws.cell(row=r, column=4, value=p['faturamento'])
            fcell.number_format = 'R$ #,##0.00'
            ws.cell(row=r, column=5, value=p['n_pedidos'])
            pcell = ws.cell(row=r, column=6, value=(p.get('pct_faturamento') or 0) / 100.0)
            pcell.number_format = '0.0%'
            ws.cell(row=r, column=7,
                    value=_estado_map_label(p.get('estado_map'), p.get('mapeado_para')))
            r += 1
        for i, w in enumerate(larguras, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()
    # Aba Consolidado primeiro (reusa a ws default).
    ws0 = wb.active
    ws0.title = _titulo_aba('Consolidado')
    _escrever_aba(ws0, dados.get('consolidado', []), 'Consolidado (todas as lojas)',
                  totais={'pedidos': dados.get('total_pedidos', 0),
                          'itens': dados.get('total_itens_vendidos', 0),
                          'faturamento': dados.get('faturamento_total', 0.0)})
    for loja in dados.get('lojas', []):
        ws = wb.create_sheet(title=_titulo_aba(loja['loja']))
        _escrever_aba(ws, loja['produtos'], loja['loja'],
                      totais={'pedidos': loja['total_pedidos'],
                              'itens': loja['total_itens'],
                              'faturamento': loja['faturamento']})

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _resolver_nome_item(tipo, item_id):
    if tipo == 'receita':
        r = Receita.query.get(item_id)
        return r.nome if r else None
    if tipo == 'produto':
        p = Produto.query.get(item_id)
        return p.nome if p else None
    if tipo == 'mp':
        from app.models import MateriaPrima
        m = MateriaPrima.query.get(item_id)
        return f'{m.nome} (MP)' if m else None
    return None


def agregar_itens_consolidado(data_inicial, data_final):
    """Versao consolidada Seru + loja propria pra uso do copilot/tools.

    Soma vendas Seru (PDV) + loja propria (PedidoOnline). VNDA APOSENTADO em
    24/06/2026 — nao soma mais nada de VNDA. Chaves `qtd_vnda`/
    `faturamento_vnda` continuam no retorno (=0) por compat com chamadores.

    Itens vendidos so no site sem correspondente Seru aparecem como linhas
    novas (fonte='site').
    """
    from app.services import loja_online_vendas, vendas_diarias

    # Seru vem do BANCO (VendaSeruDiaria), capturando o que faltar — sem
    # re-consultar a API. Import local evita ciclo (vendas_diarias importa daqui).
    seru_data = vendas_diarias.agregar_flat(data_inicial, data_final)
    vendas_online = loja_online_vendas.vendas_por_produto(data_inicial, data_final)

    seru_por_chave = {}
    seru_orfaos = []
    for p in seru_data['produtos']:
        m = p.get('match')
        if m and m.get('id') and m.get('tipo') in ('receita', 'produto'):
            chave = (m['tipo'], m['id'])
            existing = seru_por_chave.get(chave)
            if not existing or p['qtd'] > existing['qtd']:
                seru_por_chave[chave] = p
        else:
            seru_orfaos.append(p)

    linhas = []
    chaves_seru = set()
    for chave, p in seru_por_chave.items():
        qtd_online = vendas_online.get(chave, 0)
        chaves_seru.add(chave)
        p = dict(p)  # copia rasa pra nao mutar seru_data
        p['qtd_seru'] = p['qtd']
        p['qtd_vnda'] = 0
        p['qtd_online'] = qtd_online
        p['qtd'] = p['qtd_seru'] + qtd_online
        fontes = ['seru']
        if qtd_online > 0:
            fontes.append('site')
        p['fonte'] = '+'.join(fontes)
        linhas.append(p)

    # Itens que vendem so no site, sem Seru.
    for chave, qtd_online in vendas_online.items():
        if chave in chaves_seru or qtd_online <= 0:
            continue
        tipo_v, id_v = chave
        nome = _resolver_nome_item(tipo_v, id_v)
        if not nome:
            continue
        linhas.append({
            'nome': nome,
            'sku': None,
            'qtd': qtd_online,
            'qtd_seru': 0,
            'qtd_vnda': 0,
            'qtd_online': qtd_online,
            'faturamento': 0,
            'pct_faturamento': 0,
            'n_pedidos': 0,
            'fonte': 'site',
            'estado_map': 'site_only',
            'mapeado_para': {'tipo': tipo_v, 'id': id_v, 'nome': nome},
            'match': {'tipo': tipo_v, 'id': id_v, 'nome': nome, 'kind': 'exato'},
        })

    for p in seru_orfaos:
        p = dict(p)
        p['qtd_seru'] = p['qtd']
        p['qtd_vnda'] = 0
        p['qtd_online'] = 0
        p['fonte'] = 'seru'
        linhas.append(p)

    linhas.sort(key=lambda x: -x['qtd'])

    # Faturamento do site = so loja propria (PedidoOnline). VNDA aposentado.
    try:
        fat_online = loja_online_vendas.faturamento_por_dia(
            data_inicial, data_final)['total']
    except Exception:  # noqa: BLE001
        fat_online = 0.0

    fat_seru = seru_data['faturamento_total'] or 0
    return {
        'inicio': data_inicial.isoformat(),
        'fim': data_final.isoformat(),
        'total_pedidos_seru': seru_data['total_pedidos'],
        'faturamento_total': round(fat_seru + fat_online, 2),
        'faturamento_seru': round(fat_seru, 2),
        'faturamento_vnda': 0.0,
        'faturamento_online': round(fat_online, 2),
        'faturamento_fonte': 'seru+site' if fat_online > 0 else 'seru_apenas',
        'produtos': linhas,
        'vnda_aviso': 'VNDA aposentado em 06/2026',
        'lojas_no_intervalo': seru_data['lojas_no_intervalo'],
    }


def vendas_vnda_loja(data_inicial, data_final):
    """Vendas do SITE (loja propria, PedidoOnline) por produto, no formato que
    o `consultar_vendas_itens` do copilot espera.

    Usado quando o usuario filtra a venda pela loja do site (que nao tem PDV
    Seru). Desde o cutover (22/06/2026) a fonte e o `PedidoOnline` — o VNDA
    foi desligado. Nome mantido por compat do chamador (copilot). Qty por
    produto e faturamento (sem frete, por data de venda) vem de
    `loja_online_vendas`.
    """
    from app.services import loja_online_vendas

    base = {
        'inicio': data_inicial.isoformat(),
        'fim': data_final.isoformat(),
        'total_pedidos': 0,
        'faturamento_total': 0.0,
        'faturamento_fonte': 'site',
        'produtos': [],
        'vnda_aviso': None,
        'lojas_no_intervalo': [],
    }

    vd = loja_online_vendas.produtos_vendidos(data_inicial, data_final)
    try:
        fat = loja_online_vendas.faturamento_por_dia(
            data_inicial, data_final)['total']
    except Exception:  # noqa: BLE001
        fat = 0.0

    produtos = []
    for vp in vd.get('produtos', []):
        match = {'tipo': vp['tipo'], 'id': vp['id'],
                 'nome': vp['nome'], 'kind': 'exato'}
        produtos.append({
            'nome': vp['nome'], 'sku': None,
            'qtd': vp['qtd'], 'qtd_seru': 0, 'qtd_vnda': 0,
            'qtd_online': vp['qtd'],
            'faturamento': 0, 'fonte': 'site', 'match': match,
        })

    base['total_pedidos'] = vd.get('total_pedidos', 0)
    base['faturamento_total'] = round(fat, 2)
    base['produtos'] = produtos
    return base
